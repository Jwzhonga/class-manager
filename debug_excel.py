"""Debug Excel parsing（M13：文件路径和 sheet 名支持命令行参数）
用法: python3 debug_excel.py [文件路径] [sheet名]
"""
import sys
sys.path.insert(0, '.')
from app import app, db, Schedule, ScheduleImage, DAY_NAMES, MAX_PERIOD
import openpyxl, re

file_path = sys.argv[1] if len(sys.argv) > 1 else 'static/uploads/schedule_excel_20260713101351_3450.xlsx'
sheet_name = sys.argv[2] if len(sys.argv) > 2 else None

# Read all rows
wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
print(f'Sheets: {wb.sheetnames}')
ws = wb[sheet_name] if sheet_name else wb.active
print(f'Sheet: {ws.title}, rows={ws.max_row}, cols={ws.max_column}')

print('\n=== ALL CELLS ===')
for r, row in enumerate(ws.iter_rows(values_only=True), 1):
    vals = [str(c).strip() if c is not None else '' for c in row]
    print(f'R{r}: {vals}')

wb.close()
