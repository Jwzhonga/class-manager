# -*- coding: utf-8 -*-
"""
班级管理系统 - Flask Web Application
新能源班学生管理、考勤、成绩、实训、课表综合管理系统
"""

import os
import io
import csv
import re
import json
import random
import sys
from datetime import datetime, date, timedelta
from flask import (Flask, render_template, request, redirect, url_for,
                   flash, send_file, jsonify, session, Response)
from flask_sqlalchemy import SQLAlchemy

# fnOS: 将 py_packages 加入 Python 路径（install_callback 安装到此目录）
_fn_pkg = os.path.join(os.path.abspath(os.path.dirname(__file__)), 'py_packages')
if os.path.exists(_fn_pkg):
    sys.path.insert(0, _fn_pkg)
from werkzeug.security import generate_password_hash, check_password_hash
from functools import wraps
from cryptography.fernet import Fernet
from base64 import urlsafe_b64encode
import hashlib

# ── App配置 ──
app = Flask(__name__)
app.secret_key = 'class_management_secret_key_2024'
import os
# 数据库路径：fnOS 上使用 TRIM_PKGVAR（持久化数据目录），否则用默认 instance/
trim_pkgvar = os.environ.get('TRIM_PKGVAR', '')
if trim_pkgvar:
    db_path = os.path.join(trim_pkgvar, 'class_manager.db')
    app.config['SQLALCHEMY_DATABASE_URI'] = f'sqlite:///{db_path}'
    os.makedirs(os.path.dirname(db_path), exist_ok=True)
else:
    app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///master.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB
app.config['SEND_FILE_MAX_AGE_DEFAULT'] = 365 * 24 * 3600  # 静态文件缓存1年

# ── 多用户独立数据库 ──
BASE_DIR = os.path.abspath(os.path.dirname(__file__))
MASTER_DB_URI = 'sqlite:///master.db'
MASTER_DB_PATH = os.path.join(BASE_DIR, 'instance', 'master.db')
USER_DB_DIR = os.path.join(BASE_DIR, 'instance', 'users')

# gzip 压缩（跳过文件下载）
@app.after_request
def gzip_response(response):
    try:
        cl = response.content_length or 0
        if cl > 1000 and cl < 50 * 1024 * 1024:
            accept = request.headers.get('Accept-Encoding', '')
            if 'gzip' in accept:
                ct = (response.headers.get('Content-Type', '') or '')
                if 'application/octet-stream' not in ct and 'image/' not in ct:
                    import gzip as gz
                    response.direct_passthrough = False
                    data = response.get_data()
                    if isinstance(data, str):
                        data = data.encode('utf-8')
                    response.set_data(gz.compress(data))
                    response.headers['Content-Encoding'] = 'gzip'
    except:
        pass
    return response

import json as json_module
@app.template_filter('from_json')
def from_json_filter(s):
    try: return json_module.loads(s) if s else []
    except: return []

def popcount(n):
    return bin(n).count('1') if n else 0

def add_date(d, days):
    from datetime import timedelta
    return d + timedelta(days=days)

app.jinja_env.globals['now'] = lambda: datetime.now()
app.jinja_env.filters['popcount'] = popcount

UPLOAD_FOLDER = os.path.join(BASE_DIR, 'static', 'uploads')
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# ── 登录验证装饰器 ──
def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            flash('请先登录')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated


# 全局登录检查（白名单路由除外）
PUBLIC_ROUTES = {'login', 'register', 'logout', 'static'}

@app.before_request
def check_login():
    if request.endpoint and request.endpoint not in PUBLIC_ROUTES and 'user_id' not in session:
        flash('请先登录')
        return redirect(url_for('login'))
    # 每次请求开始时，确保使用正确的数据库
    if request.endpoint and 'user_id' in session and request.endpoint not in PUBLIC_ROUTES:
        # 已登录用户 → 切换到用户自己的业务数据库
        switch_to_user_db(session['user_id'])
    elif request.endpoint and request.endpoint in PUBLIC_ROUTES:
        # 公开页面 → 确保连接主数据库
        if not is_on_master_db():
            switch_db(MASTER_DB_URI)


def get_current_semester_id():
    """获取当前选中的学期ID"""
    sem_id = session.get('semester_id')
    if sem_id:
        return sem_id
    current = Semester.query.filter_by(is_current=True).first()
    return current.id if current else None


def get_semester_students():
    """获取当前学期的学生列表"""
    sem_id = get_current_semester_id()
    q = Student.query
    if sem_id:
        q = q.filter_by(semester_id=sem_id)
    return q.order_by(Student.id).all()


def get_semester_projects():
    """获取当前学期的实训项目"""
    sem_id = get_current_semester_id()
    q = TrainingProject.query
    if sem_id:
        q = q.filter_by(semester_id=sem_id)
    return q.order_by(TrainingProject.id).all()


def get_semester_schedule():
    """获取当前学期的课表"""
    sem_id = get_current_semester_id()
    q = Schedule.query
    if sem_id:
        q = q.filter_by(semester_id=sem_id)
    return q.all()


def get_semester_subject_objects():
    """获取当前学期的Subject对象列表"""
    sem_id = get_current_semester_id()
    q = Subject.query
    if sem_id:
        q = q.filter_by(semester_id=sem_id)
    return q.order_by(Subject.id).all()


# 学期上下文注入
@app.context_processor
def inject_semester():
    semesters = Semester.query.order_by(Semester.start_date.desc()).all()
    current = Semester.query.filter_by(is_current=True).first()
    if not current and semesters:
        current = semesters[0]
    sem_id = session.get('semester_id') or (current.id if current else None)
    selected = Semester.query.get(sem_id) if sem_id else current
    return dict(semesters=semesters, current_semester=selected)


TRUANCY_WARN_HOURS = 80  # 旷课预警阈值

db = SQLAlchemy(app)

# ── 多用户独立数据库函数（实际实现） ──
def switch_db(database_uri):
    """切换Flask-SQLAlchemy到指定的数据库URI（关闭旧连接，重新创建引擎）"""
    db.session.remove()
    app.config['SQLALCHEMY_DATABASE_URI'] = database_uri
    # 重新初始化引擎
    from sqlalchemy import create_engine
    abs_uri = _resolve_db_uri(database_uri)
    new_engine = create_engine(abs_uri)
    try:
        ext = app.extensions['sqlalchemy']
        for ref in list(ext._app_engines.data.keys()):
            if ref() is app:
                old = ext._app_engines.data[ref][None]
                old.dispose()
                ext._app_engines.data[ref][None] = new_engine
                break
    except:
        pass
    # 应用所有迁移
    _run_db_migrations()
    db.create_all()


def _resolve_db_uri(uri):
    """将相对路径的sqlite URI转为绝对路径"""
    if uri == MASTER_DB_URI:
        return f'sqlite:///{MASTER_DB_PATH}'
    if uri.startswith('sqlite:///') and not uri.startswith('sqlite:////'):
        rel = uri[10:]
        path = os.path.join(BASE_DIR, 'instance', rel)
        return f'sqlite:///{path}'
    return uri


def _get_db_path():
    """获取当前用户的数据库文件路径"""
    try:
        uid = session.get('user_id')
        if uid:
            return os.path.join(USER_DB_DIR, f'u{uid}.db')
    except:
        pass
    return MASTER_DB_PATH


def _run_db_migrations():
    """对新创建的数据库执行ALTER TABLE迁移"""
    db_path = _get_db_path()
    if not os.path.exists(db_path):
        return
    try:
        conn = sqlite3.connect(db_path)
        c = conn.cursor()
        # student表迁移
        for col, typ in [('dormitory','VARCHAR(32)'),('special_family','VARCHAR(64)'),
                         ('special_family_note','TEXT'),('special_physical','VARCHAR(8)'),
                         ('special_physical_note','TEXT'),('remark','TEXT')]:
            try: c.execute(f'ALTER TABLE student ADD COLUMN {col} {typ} DEFAULT ""')
            except: pass
        for col, typ in [('status','VARCHAR(16)')]:
            try: c.execute(f'ALTER TABLE student ADD COLUMN {col} {typ} DEFAULT "active"')
            except: pass
        for col, typ in [('withdrawn_reason','TEXT')]:
            try: c.execute(f'ALTER TABLE student ADD COLUMN {col} {typ} DEFAULT ""')
            except: pass
        try: c.execute('ALTER TABLE subject ADD COLUMN class_name VARCHAR(64) DEFAULT ""')
        except: pass
        conn.commit()
        conn.close()
    except:
        pass


def get_master_conn():
    """获取主数据库连接（raw sqlite3，不依赖Flask-SQLAlchemy引擎切换）"""
    os.makedirs(os.path.join(BASE_DIR, 'instance'), exist_ok=True)
    conn = sqlite3.connect(MASTER_DB_PATH)
    conn.row_factory = sqlite3.Row
    # 确保master_user表存在
    c = conn.cursor()
    c.execute('''CREATE TABLE IF NOT EXISTS master_user (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        username TEXT UNIQUE NOT NULL,
        password_hash TEXT NOT NULL,
        db_name TEXT DEFAULT '',
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )''')
    conn.commit()
    return conn

def switch_to_user_db(user_id):
    """切换到指定用户的独立业务数据库，如不存在则自动创建表"""
    user_db_dir = os.path.join(BASE_DIR, 'instance', 'users')
    os.makedirs(user_db_dir, exist_ok=True)
    db_path = os.path.join(user_db_dir, f'u{user_id}.db')
    switch_db(f'sqlite:///{db_path}')

def is_on_master_db():
    """判断当前是否连接的是主数据库"""
    uri = app.config.get('SQLALCHEMY_DATABASE_URI', '')
    return MASTER_DB_URI in uri or MASTER_DB_PATH in uri

# ── 常数 ──
DAY_NAMES = ['周一', '周二', '周三', '周四', '周五']
SCORE_TYPES = ['课堂表现', '作业质量', '课堂笔记', '考试成绩']
SCORE_WEIGHTS = [0.2, 0.2, 0.2, 0.4]  # 各项权重

def calc_expiry(incident_date):
    """根据预设计算处分到期日"""
    from datetime import timedelta
    preset = request.form.get('expiry_preset', '')
    custom = request.form.get('expiry_date', '')
    if preset == 'custom' and custom:
        try: return date.fromisoformat(custom)
        except: return None
    elif preset == '1m':
        d = incident_date.replace(month=incident_date.month + 1) if incident_date.month < 12 else date(incident_date.year + 1, 1, incident_date.day)
        return d
    elif preset == '3m':
        m = incident_date.month + 3
        y = incident_date.year + (m - 1) // 12
        m = ((m - 1) % 12) + 1
        return date(y, m, incident_date.day)
    elif preset == '6m':
        m = incident_date.month + 6
        y = incident_date.year + (m - 1) // 12
        m = ((m - 1) % 12) + 1
        return date(y, m, incident_date.day)
    elif preset == '1y':
        return date(incident_date.year + 1, incident_date.month, incident_date.day)
    return None


# 科目管理辅助函数
def get_subjects():
    """获取当前学期的科目列表"""
    sem_id = get_current_semester_id()
    q = Subject.query
    if sem_id:
        q = q.filter_by(semester_id=sem_id)
    return [s.name for s in q.order_by(Subject.id).all()]


class MasterUser(db.Model):
    """主数据库用户模型 — 仅存于 master.db，只保存登录凭据"""
    __tablename__ = 'master_user'
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(32), unique=True, nullable=False)
    password_hash = db.Column(db.String(256), nullable=False)
    db_name = db.Column(db.String(64), default='')  # 对应的用户数据库文件名
    created_at = db.Column(db.DateTime, default=datetime.now)

    def set_password(self, pw):
        self.password_hash = generate_password_hash(pw, method='pbkdf2:sha256')

    def check_password(self, pw):
        return check_password_hash(self.password_hash, pw)


class Semester(db.Model):
    __tablename__ = 'semester'
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(64), nullable=False)
    start_date = db.Column(db.Date, nullable=False)
    end_date = db.Column(db.Date, nullable=False)
    is_current = db.Column(db.Boolean, default=False)
    created_at = db.Column(db.DateTime, default=datetime.now)


ATTENDANCE_STATUS_NAMES = {
    'present': '出勤', 'late': '迟到', 'sick': '病假',
    'personal': '事假', 'truant': '旷课'
}
MAX_PERIOD = 7  # 每天最多7节课
PERIOD_TIMES = {
    1: '8:10-8:55', 2: '9:05-9:50', 3: '10:20-11:05',
    4: '11:15-12:00', 5: '14:25-15:10', 6: '15:20-16:05', 7: '17:00-17:45',
    8: '夜自习1', 9: '夜自习2', 10: '夜自习3'
}


# ══════════════════════════════════════════════
# 模型定义
# ══════════════════════════════════════════════

class Student(db.Model):
    __tablename__ = 'student'
    id = db.Column(db.Integer, primary_key=True)
    semester_id = db.Column(db.Integer, db.ForeignKey('semester.id'), nullable=True)
    name = db.Column(db.String(64), nullable=False)
    student_id = db.Column(db.String(20), unique=False, nullable=False)
    class_name = db.Column(db.String(64), default='')
    gender = db.Column(db.String(8), default='')
    id_card = db.Column(db.String(32), default='')
    phone = db.Column(db.String(32), default='')
    parent_name = db.Column(db.String(64), default='')
    parent_phone = db.Column(db.String(32), default='')
    parent_relation = db.Column(db.String(16), default='')
    ethnic = db.Column(db.String(16), default='')
    political_status = db.Column(db.String(16), default='')
    graduate_school = db.Column(db.String(128), default='')
    residence_type = db.Column(db.String(32), default='')
    live_mode = db.Column(db.String(16), default='')
    poverty_status = db.Column(db.String(16), default='')
    address = db.Column(db.String(256), default='')
    major = db.Column(db.String(64), default='')
    notes = db.Column(db.Text, default='')
    dormitory = db.Column(db.String(32), default='')
    special_family = db.Column(db.String(64), default='')
    special_family_note = db.Column(db.Text, default='')
    special_physical = db.Column(db.String(8), default='')
    special_physical_note = db.Column(db.Text, default='')
    remark = db.Column(db.Text, default='')
    status = db.Column(db.String(16), default='active')  # active / withdrawn
    withdrawn_reason = db.Column(db.Text, default='')
    created_at = db.Column(db.DateTime, default=datetime.now)

    def to_dict(self):
        return {c.name: getattr(self, c.name) for c in self.__table__.columns}


class Attendance(db.Model):
    __tablename__ = 'attendance'
    id = db.Column(db.Integer, primary_key=True)
    student_id = db.Column(db.Integer, db.ForeignKey('student.id'), nullable=False)
    date = db.Column(db.Date, nullable=False)
    status = db.Column(db.String(16), default='present')
    reason = db.Column(db.String(256), default='')
    course_name = db.Column(db.String(64), default='')
    period = db.Column(db.Integer, default=1)
    image_path = db.Column(db.String(256), default='')
    semester_id = db.Column(db.Integer, db.ForeignKey('semester.id'), nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.now)

    student = db.relationship('Student', backref=db.backref('attendance_records', lazy='dynamic'))


class Grade(db.Model):
    __tablename__ = 'grade'
    id = db.Column(db.Integer, primary_key=True)
    student_id = db.Column(db.Integer, db.ForeignKey('student.id'), nullable=False)
    subject = db.Column(db.String(64), nullable=False)  # 科目名称
    performance_score = db.Column(db.Float, default=0)  # 课堂表现
    homework_score = db.Column(db.Float, default=0)    # 作业质量
    notes_score = db.Column(db.Float, default=0)       # 课堂笔记
    exam_score = db.Column(db.Float, default=0)         # 考试成绩
    comprehensive_score = db.Column(db.Float, default=0)
    overall_score = db.Column(db.Float, default=0)
    semester_id = db.Column(db.Integer, db.ForeignKey('semester.id'), nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.now)
    student = db.relationship('Student', backref=db.backref('grades', lazy='dynamic'))

    def calc_comprehensive(self):
        """计算学科综合成绩 = 课堂表现×20% + 作业质量×20% + 课堂笔记×20% + 考试成绩×40%"""
        self.comprehensive_score = round(
            self.performance_score * 0.2 +
            self.homework_score * 0.2 +
            self.notes_score * 0.2 +
            self.exam_score * 0.4, 1)
        return self.comprehensive_score


class Subject(db.Model):
    __tablename__ = 'subject'
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(64), unique=False, nullable=False)
    teacher = db.Column(db.String(32), default='')
    class_name = db.Column(db.String(64), default='')  # 班级名称
    semester_id = db.Column(db.Integer, db.ForeignKey('semester.id'), nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.now)


class CourseStudent(db.Model):
    """任课管理 - 科目下的学生名单（独立于学生管理中的Student表）"""
    __tablename__ = 'course_student'
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(64), nullable=False)
    course_id = db.Column(db.Integer, db.ForeignKey('subject.id'), nullable=False)
    semester_id = db.Column(db.Integer, db.ForeignKey('semester.id'), nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.now)

    course = db.relationship('Subject', backref=db.backref('course_students', lazy='dynamic'))


class Discipline(db.Model):
    __tablename__ = 'discipline'
    id = db.Column(db.Integer, primary_key=True)
    student_id = db.Column(db.Integer, db.ForeignKey('student.id'), nullable=False)
    incident_date = db.Column(db.Date, nullable=False, default=date.today)  # 何时
    location = db.Column(db.String(128), default='')  # 何地
    reason = db.Column(db.Text, default='')  # 何事
    punishment = db.Column(db.String(128), default='')  # 何种处分
    expiry_date = db.Column(db.Date, nullable=True)  # 处分期限
    notes = db.Column(db.Text, default='')
    image_notification = db.Column(db.Text, default='')  # JSON数组
    image_letter = db.Column(db.Text, default='')  # JSON数组
    semester_id = db.Column(db.Integer, db.ForeignKey('semester.id'), nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.now)
    student = db.relationship('Student', backref=db.backref('disciplines', lazy='dynamic'))


class ViolationRecord(db.Model):
    __tablename__ = 'violation_record'
    id = db.Column(db.Integer, primary_key=True)
    student_id = db.Column(db.Integer, db.ForeignKey('student.id'), nullable=False)
    date = db.Column(db.Date, nullable=False, default=date.today)
    reason = db.Column(db.Text, default='')
    reflection_days = db.Column(db.Integer, default=1)
    reflection_start = db.Column(db.Date, nullable=True)
    reflection_end = db.Column(db.Date, nullable=True)
    notes = db.Column(db.Text, default='')
    semester_id = db.Column(db.Integer, db.ForeignKey('semester.id'), nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.now)
    student = db.relationship('Student', backref=db.backref('violations', lazy='dynamic'))


class ClassFund(db.Model):
    __tablename__ = 'class_fund'
    id = db.Column(db.Integer, primary_key=True)
    date = db.Column(db.Date, nullable=False, default=date.today)
    type = db.Column(db.String(8), nullable=False)  # income / expense
    amount = db.Column(db.Float, nullable=False, default=0)
    reason = db.Column(db.Text, default='')
    voucher = db.Column(db.String(256), default='')  # 凭据图片路径
    notes = db.Column(db.Text, default='')
    semester_id = db.Column(db.Integer, db.ForeignKey('semester.id'), nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.now)


class TrainingProject(db.Model):
    __tablename__ = 'training_project'
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(128), nullable=False)
    description = db.Column(db.Text, default='')
    max_score = db.Column(db.Float, default=100)
    category = db.Column(db.String(64), default='')
    semester_id = db.Column(db.Integer, db.ForeignKey('semester.id'), nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.now)


# 实训分组-学生 多对多关联表
training_group_student = db.Table('training_group_student',
    db.Column('group_id', db.Integer, db.ForeignKey('training_group.id'), primary_key=True),
    db.Column('student_id', db.Integer, db.ForeignKey('student.id'), primary_key=True)
)


class TrainingGroup(db.Model):
    __tablename__ = 'training_group'
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(128), nullable=False)
    semester_id = db.Column(db.Integer, db.ForeignKey('semester.id'), nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.now)

    students = db.relationship('Student', secondary=training_group_student,
                               backref=db.backref('training_groups', lazy='dynamic'),
                               lazy='dynamic')
    records = db.relationship('TrainingRecord', backref=db.backref('group', lazy='joined'),
                              lazy='dynamic')


class TrainingRecord(db.Model):
    __tablename__ = 'training_record'
    id = db.Column(db.Integer, primary_key=True)
    student_id = db.Column(db.Integer, db.ForeignKey('student.id'), nullable=False)
    project_id = db.Column(db.Integer, db.ForeignKey('training_project.id'), nullable=False)
    group_id = db.Column(db.Integer, db.ForeignKey('training_group.id'), nullable=True)
    score = db.Column(db.Float, default=0)
    quality_notes = db.Column(db.String(64), default='')
    completion_date = db.Column(db.Date, nullable=True)
    instructor_notes = db.Column(db.Text, default='')
    semester_id = db.Column(db.Integer, db.ForeignKey('semester.id'), nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.now)

    student = db.relationship('Student', backref=db.backref('training_records', lazy='dynamic'))
    project = db.relationship('TrainingProject', backref=db.backref('records', lazy='dynamic'))


class Schedule(db.Model):
    __tablename__ = 'schedule'
    id = db.Column(db.Integer, primary_key=True)
    day_of_week = db.Column(db.Integer, nullable=False)  # 0=周一 ... 4=周五
    period = db.Column(db.Integer, nullable=False)       # 1-8
    course_name = db.Column(db.String(128), nullable=False)
    teacher = db.Column(db.String(64), default='')
    location = db.Column(db.String(128), default='')
    is_training = db.Column(db.Boolean, default=False)
    semester_id = db.Column(db.Integer, db.ForeignKey('semester.id'), nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.now)


class ScheduleImage(db.Model):
    __tablename__ = 'schedule_image'
    id = db.Column(db.Integer, primary_key=True)
    image_path = db.Column(db.String(256), nullable=False)
    raw_text = db.Column(db.Text, default='')
    ocr_result = db.Column(db.Text, default='')  # JSON string
    created_at = db.Column(db.DateTime, default=datetime.now)


class Seat(db.Model):
    __tablename__ = 'seat'
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(64), nullable=False)  # 座位名称如"第1排第1列"
    row_num = db.Column(db.Integer, nullable=False)
    col_num = db.Column(db.Integer, nullable=False)
    semester_id = db.Column(db.Integer, db.ForeignKey('semester.id'), nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.now)


class SeatAssignment(db.Model):
    __tablename__ = 'seat_assignment'
    id = db.Column(db.Integer, primary_key=True)
    seat_id = db.Column(db.Integer, db.ForeignKey('seat.id'), nullable=False)
    student_id = db.Column(db.Integer, db.ForeignKey('student.id'), nullable=False)
    semester_id = db.Column(db.Integer, db.ForeignKey('semester.id'), nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.now)

    seat = db.relationship('Seat', backref=db.backref('assignments', lazy='dynamic'))
    student = db.relationship('Student', backref=db.backref('seat_assignments', lazy='dynamic'))


# 创建所有表（初始化为主数据库 - master.db）
with app.app_context():
    # 仅初始化 master.db 的 master_user 表
    import sqlite3
    master_db_dir = os.path.join(BASE_DIR, 'instance')
    os.makedirs(master_db_dir, exist_ok=True)
    master_db_path = os.path.join(master_db_dir, 'master.db')
    master_conn = sqlite3.connect(master_db_path)
    master_c = master_conn.cursor()
    master_c.execute('''CREATE TABLE IF NOT EXISTS master_user (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        username VARCHAR(32) UNIQUE NOT NULL,
        password_hash VARCHAR(256) NOT NULL,
        db_name VARCHAR(64) DEFAULT '',
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )''')
    master_conn.commit()
    master_conn.close()
    # 设置当前数据库连接为master.db
    db.session.remove()
    app.config['SQLALCHEMY_DATABASE_URI'] = f'sqlite:///master.db'
    if db.engine:
        db.engine.dispose()


# ══════════════════════════════════════════════
# 辅助函数
# ══════════════════════════════════════════════

def generate_student_id():
    """自动生成学号 STU001, STU002 ..."""
    last = Student.query.order_by(Student.id.desc()).first()
    if last and last.student_id and last.student_id.startswith('STU'):
        try:
            num = int(last.student_id[3:]) + 1
            return f'STU{num:03d}'
        except ValueError:
            pass
    return 'STU001'


def get_age(birth_id_card):
    """从身份证号推算年龄"""
    if not birth_id_card or len(birth_id_card) < 18:
        return None
    try:
        birth_year = int(birth_id_card[6:10])
        birth_month = int(birth_id_card[10:12])
        birth_day = int(birth_id_card[12:14])
        today = date.today()
        age = today.year - birth_year
        if (today.month, today.day) < (birth_month, birth_day):
            age -= 1
        return age
    except (ValueError, IndexError):
        return None


def get_region_from_idcard(id_card):
    """从身份证号提取省市代码"""
    if not id_card or len(id_card) < 6:
        return None
    code = id_card[:6]
    # 常见甘肃代码映射
    region_map = {
        '6201': '兰州', '6202': '嘉峪关', '6203': '金昌', '6204': '白银',
        '6205': '天水', '6206': '武威', '6207': '张掖', '6208': '平凉',
        '6209': '酒泉', '6210': '庆阳', '6211': '定西', '6212': '陇南',
        '6229': '临夏', '6230': '甘南',
        '620121': '永登', '620122': '皋兰', '620123': '榆中',
    }
    # 先匹配最长
    for c in sorted(region_map.keys(), key=len, reverse=True):
        if code.startswith(c):
            return region_map[c]
    return '其他'


def parse_excel_schedule(file_path, sheet_name=None):
    """从Excel课表直接读取内容"""
    courses = []
    try:
        if file_path.endswith('.xls'):
            import xlrd
            wb = xlrd.open_workbook(file_path)
            sheet = wb.sheet_by_name(sheet_name) if sheet_name else wb.sheet_by_index(0)
            rows = [[str(sheet.cell_value(r, c)).strip() for c in range(sheet.ncols)] for r in range(sheet.nrows)]
        else:
            from openpyxl import load_workbook
            wb = load_workbook(file_path, read_only=True, data_only=True)
            ws = wb[sheet_name] if sheet_name else wb.active
            rows = []
            for row in ws.iter_rows(values_only=True):
                rows.append([str(c).strip() if c is not None else '' for c in row])
            wb.close()
        
        # 找表头行：包含"星期一"或"周一"
        header_idx = None
        for idx, row in enumerate(rows):
            text = '|'.join(row)
            if '星期一' in text or '星期' in text:
                header_idx = idx
                break
        
        if header_idx is None:
            flash('未找到课表表头，请检查Excel是否有"星期一~星期五"列')
            return courses
        
        # 映射星期列
        day_names = [('星期一',0),('星期二',1),('星期三',2),('星期四',3),('星期五',4),('周一',0),('周二',1),('周三',2),('周四',3),('周五',4)]
        day_col = {}
        header = rows[header_idx]
        for ci, cell in enumerate(header):
            for kw, di in day_names:
                if kw in cell:
                    day_col[di] = ci
                    break
        
        if not day_col:
            # 默认列1-5
            day_col = {i: i+1 for i in range(5)}
        
        # 数据从表头下一行开始
        for ri in range(header_idx + 1, len(rows)):
            row = rows[ri]
            if not row or not row[0]:
                continue
            first = row[0].strip()
            
            # 跳过底部备注行（如"2026年2月23日"）
            if '年' in first and '月' in first:
                continue
            if '班主任' in first:
                continue
            
            # 提取节次号："上午1"→1、"下午2"→6 等
            period = None
            am_match = re.search(r'上午(\d+)', first)
            pm_match = re.search(r'下午(\d+)', first)
            if am_match:
                period = int(am_match.group(1))
            elif pm_match:
                period = int(pm_match.group(1)) + 4  # 下午1→5, 下午2→6, 下午3→7
            else:
                m = re.search(r'(\d+)', first)
                if m:
                    period = int(m.group(1))
                else:
                    continue
            if period is None or period < 1 or period > MAX_PERIOD:
                continue
            
            for di, ci in day_col.items():
                if ci < len(row):
                    cell = row[ci].strip()
                    if cell and cell not in ('-', '—', '–', '', 'None', '/', '\\'):
                        # 单元格格式："课程名\\n教师名" 或 "课程名"
                        parts = cell.split('\n')
                        course_name = parts[0].strip()
                        teacher = parts[1].strip() if len(parts) >= 2 else ''
                        if not course_name:
                            continue
                        courses.append({
                            'day_of_week': di,
                            'period': period,
                            'course_name': course_name[:64],
                            'teacher': teacher[:32],
                            'location': '',
                            'is_training': any(kw in course_name for kw in ['检修', '构造', '驱动', '电气', '实训'])
                        })
        
        if not courses:
            flash('未能从Excel中解析到课程。确认第1列有节次号（如"上午1"、"1"），第2-6列为星期一~星期五')
        return courses
        
    except Exception as e:
        flash(f'Excel解析失败: {e}')
        return courses
# ══════════════════════════════════════════════
# ── 认证管理 ──
# ══════════════════════════════════════════════

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        # 切换到主数据库验证用户身份
        if not is_on_master_db():
            switch_db(MASTER_DB_URI)
        user = MasterUser.query.filter_by(username=username).first()
        if user and user.check_password(password):
            session['user_id'] = user.id
            session['username'] = user.username
            # 登录成功后切换到用户自己的业务数据库
            switch_to_user_db(user.id)
            flash(f'欢迎回来，{user.username}')
            return redirect(url_for('index'))
        flash('用户名或密码错误')
    return render_template('login.html')


@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        if not username or not password:
            flash('请填写所有字段')
            return render_template('register.html')
        # 切换到主数据库检查是否已存在
        if not is_on_master_db():
            switch_db(MASTER_DB_URI)
        if MasterUser.query.filter_by(username=username).first():
            flash('用户名已存在')
            return render_template('register.html')
        # 在主数据库创建用户
        user = MasterUser(username=username)
        user.set_password(password)
        db.session.add(user)
        db.session.commit()
        # 为新用户创建独立的业务数据库并初始化
        switch_to_user_db(user.id)
        # 创建默认学期
        from datetime import date
        s = Semester(name='2025-2026学年度第1学期', start_date=date(2025, 9, 1),
                     end_date=date(2026, 1, 15), is_current=True)
        db.session.add(s)
        db.session.commit()
        session['semester_id'] = s.id
        flash(f'注册成功，请登录')
        return redirect(url_for('login'))
    return render_template('register.html')


@app.route('/logout')
def logout():
    session.clear()
    flash('已退出登录')
    return redirect(url_for('login'))


@app.route('/semester/set', methods=['POST'])
def semester_set():
    sem_id = request.form.get('semester_id', type=int)
    if sem_id and Semester.query.get(sem_id):
        session['semester_id'] = sem_id
    return redirect(request.referrer or url_for('index'))


@app.route('/semester/add', methods=['POST'])
def semester_add():
    start_year = request.form.get('start_year', '').strip()
    sem_num = request.form.get('semester_num', '1')
    start = request.form.get('start_date', '')
    end = request.form.get('end_date', '')
    if start_year and start and end:
        try:
            end_year = int(start_year) + 1
            name = f'{start_year}-{end_year}学年度第{sem_num}学期'
            s = Semester(name=name, start_date=date.fromisoformat(start),
                         end_date=date.fromisoformat(end), is_current=False)
            db.session.add(s)
            db.session.commit()
            session['semester_id'] = s.id
            flash(f'学期「{name}」已创建')
        except:
            flash('日期格式错误')
    else:
        flash('请填写完整信息')
    return redirect(url_for('index'))


@app.route('/students/quick-edit', methods=['POST'])
@login_required
def student_quick_edit():
    """快速编辑学生字段（下拉切换）"""
    sid = request.form.get('id', type=int)
    field = request.form.get('field', '')
    value = request.form.get('value', '')
    student = Student.query.get_or_404(sid)
    if hasattr(student, field):
        setattr(student, field, value)
        db.session.commit()
        return {'ok': True, 'value': value}
    return {'ok': False}, 400


@app.route('/students/batch-delete', methods=['POST'])
@login_required
def student_batch_delete():
    """批量删除学生"""
    ids = request.form.getlist('ids')
    if not ids:
        flash('未选择学生')
        return redirect(url_for('student_list'))
    count = 0
    action = request.form.get('action', 'delete')
    if action == 'withdraw':
        for sid in ids:
            s = Student.query.get(int(sid))
            reason = request.form.get(f'reason_{sid}', '').strip()
            if s and s.status != 'withdrawn' and reason:
                s.status = 'withdrawn'
                s.withdrawn_reason = reason
                db.session.add(s)
                count += 1
        db.session.commit()
        flash(f'已将 {count} 名学生标记为流失')
    else:
        for sid in ids:
            s = Student.query.get(int(sid))
            if s:
                Attendance.query.filter_by(student_id=s.id).delete()
                Grade.query.filter_by(student_id=s.id).delete()
                TrainingRecord.query.filter_by(student_id=s.id).delete()
                Discipline.query.filter_by(student_id=s.id).delete()
                ViolationRecord.query.filter_by(student_id=s.id).delete()
                db.session.delete(s)
                count += 1
        db.session.commit()
        flash(f'已删除 {count} 名学生')
    return redirect(url_for('student_list'))


@app.route('/students/withdraw', methods=['POST'])
@login_required
def student_batch_withdraw():
    """批量标记学生为流失"""
    ids = request.form.getlist('ids')
    if not ids:
        flash('未选择学生')
        return redirect(url_for('student_list'))
    count = 0
    for sid in ids:
        s = Student.query.get(int(sid))
        if s and s.status != 'withdrawn':
            s.status = 'withdrawn'
            db.session.add(s)
            count += 1
    db.session.commit()
    flash(f'已将 {count} 名学生标记为流失')
    return redirect(url_for('student_list'))


@app.route('/semester/<int:id>/rename', methods=['POST'])
@login_required
def semester_rename(id):
    sem = Semester.query.get_or_404(id)
    name = request.form.get('name', '').strip()
    if name:
        sem.name = name
        db.session.commit()
        flash(f'学期已重命名为「{name}」')
    return redirect(request.referrer or url_for('index'))


@app.route('/semester/<int:id>/delete', methods=['POST'])
def semester_delete(id):
    sem = Semester.query.get_or_404(id)
    name = sem.name
    # 删除关联数据
    Attendance.query.filter_by(semester_id=id).delete()
    Discipline.query.filter_by(semester_id=id).delete()
    Grade.query.filter_by(semester_id=id).delete()
    TrainingRecord.query.filter_by(semester_id=id).delete()
    ViolationRecord.query.filter_by(semester_id=id).delete()
    ClassFund.query.filter_by(semester_id=id).delete()
    CourseStudent.query.filter_by(semester_id=id).delete()
    Subject.query.filter_by(semester_id=id).delete()
    TrainingProject.query.filter_by(semester_id=id).delete()
    TrainingGroup.query.filter_by(semester_id=id).delete()
    SeatAssignment.query.filter_by(semester_id=id).delete()
    Seat.query.filter_by(semester_id=id).delete()
    Schedule.query.filter_by(semester_id=id).delete()
    db.session.delete(sem)
    db.session.commit()
    if session.get('semester_id') == id:
        session.pop('semester_id', None)
    flash(f'学期「{name}」及关联数据已删除')
    return redirect(url_for('index'))


# ── 首页 - 仪表盘 ──
# ══════════════════════════════════════════════

@app.route('/')
@login_required
def index():
    all_students = get_semester_students()
    total_students = len([s for s in all_students if s.status != 'withdrawn'])
    today = date.today()
    sem_id = get_current_semester_id()

    # 今日考勤统计
    today_attendance = Attendance.query.filter_by(date=today).all()
    if sem_id:
        today_attendance = [a for a in today_attendance if a.semester_id == sem_id]
    today_late = sum(1 for a in today_attendance if a.status == 'late')
    today_sick = sum(1 for a in today_attendance if a.status == 'sick')
    today_personal = sum(1 for a in today_attendance if a.status == 'personal')
    today_truant = sum(1 for a in today_attendance if a.status == 'truant')

    # 本月旷课统计
    month_start = date(today.year, today.month, 1)
    month_q = Attendance.query.filter(Attendance.date >= month_start, Attendance.status == 'truant')
    if sem_id:
        month_q = month_q.filter_by(semester_id=sem_id)
    month_truancy = month_q.count()

    # 旷课预警
    warn_students = []
    all_truancy = []
    all_students = get_semester_students()
    for s in all_students:
        q = Attendance.query.filter_by(student_id=s.id, status='truant')
        if sem_id:
            q = q.filter_by(semester_id=sem_id)
        truant_count = sum(popcount(r.period) for r in q.all())
        if truant_count >= TRUANCY_WARN_HOURS:
            warn_students.append({
                'name': s.name,
                'student_id': s.student_id,
                'hours': truant_count
            })
        if truant_count > 0:
            all_truancy.append({
                'name': s.name,
                'hours': truant_count
            })
    # 按旷课节数降序排列
    all_truancy.sort(key=lambda x: x['hours'], reverse=True)

    # 即将到期处分（30天内）
    soon = date.today() + timedelta(days=30)
    expiring = Discipline.query.filter(
        Discipline.expiry_date.isnot(None),
        Discipline.expiry_date <= soon,
        Discipline.expiry_date >= date.today()
    )
    if sem_id:
        expiring = expiring.filter_by(semester_id=sem_id)
    expiring = expiring.order_by(Discipline.expiry_date).all()
    
    # 最近5条处分记录
    rq = Discipline.query
    if sem_id:
        rq = rq.filter_by(semester_id=sem_id)
    recent_disciplines = rq.order_by(Discipline.incident_date.desc()).limit(5).all()

    stats = {
        'total_students': total_students,
        'today': today,
        'today_late': today_late,
        'today_sick': today_sick,
        'today_personal': today_personal,
        'today_truant': today_truant,
        'month_truancy': month_truancy,
        'truancy_warn_students': warn_students,
    }
    return render_template('index.html', stats=stats, warn_hours=TRUANCY_WARN_HOURS, all_truancy=all_truancy,
                          expiring_disciplines=expiring, recent_disciplines=recent_disciplines)


# ══════════════════════════════════════════════
# 学生管理
# ══════════════════════════════════════════════

@app.route('/students')
def student_list():
    filter_type = request.args.get('filter', '')
    filter_value = request.args.get('value', '')
    sem_id = get_current_semester_id()
    students_query = Student.query.order_by(Student.id)
    if sem_id:
        students_query = students_query.filter_by(semester_id=sem_id)
    if filter_type == 'poverty':
        students_query = students_query.filter(Student.poverty_status.in_(['贫困户', '是']))
    elif filter_type == 'boarding':
        students_query = students_query.filter(Student.live_mode == '住校')
    elif filter_type == 'day':
        students_query = students_query.filter(Student.live_mode == '走读')
    elif filter_type == 'minority':
        students_query = students_query.filter(Student.ethnic.notin_(['', '汉族']))
    elif filter_type == 'agricultural':
        students_query = students_query.filter(Student.residence_type == '农业户口')
    elif filter_type == 'non_agricultural':
        students_query = students_query.filter(Student.residence_type == '非农业户口')
    elif filter_type == 'special_family':
        students_query = students_query.filter(Student.special_family != '', Student.special_family.isnot(None))
    elif filter_type == 'special_physical':
        students_query = students_query.filter(Student.special_physical == '是')
    elif filter_type == 'withdrawn':
        students_query = students_query.filter(Student.status == 'withdrawn')
    elif filter_type == 'age_range':
        all_s = Student.query.all()
        filtered = []
        for s in all_s:
            age = get_age(s.id_card)
            if age is None:
                continue
            if filter_value == '15以下' and age < 16:
                filtered.append(s)
            elif filter_value == '16-17' and age in (16, 17):
                filtered.append(s)
            elif filter_value == '18-19' and age in (18, 19):
                filtered.append(s)
            elif filter_value == '20以上' and age >= 20:
                filtered.append(s)
        students_query = filtered if filtered else Student.query.filter_by(id=0)
    elif filter_type == 'region':
        all_s = Student.query.all()
        filtered = [s for s in all_s if get_region_from_idcard(s.id_card) == filter_value]
        students_query = filtered if filtered else Student.query.filter_by(id=0)

    if isinstance(students_query, list):
        students = students_query
        display_count = len(students)
        total = len([s for s in get_semester_students() if s.status != 'withdrawn'])
    else:
        # 非过滤模式排除流失学生
        if filter_type != 'withdrawn':
            students_query = students_query.filter(Student.status != 'withdrawn')
        students = students_query.all()
        display_count = len(students)
        total = len([s for s in get_semester_students() if s.status != 'withdrawn'])

    # 统计（排除流失学生）
    all_students = [s for s in get_semester_students() if s.status != 'withdrawn']
    poverty_count = sum(1 for s in all_students if s.poverty_status in ('贫困户', '是'))
    boarding_count = sum(1 for s in all_students if s.live_mode == '住校')
    day_count = sum(1 for s in all_students if s.live_mode == '走读')
    minority_count = sum(1 for s in all_students if s.ethnic not in ('', '汉族'))
    agri_count = sum(1 for s in all_students if s.residence_type == '农业户口')
    non_agri_count = sum(1 for s in all_students if s.residence_type == '非农业户口')
    special_family_count = sum(1 for s in all_students if s.special_family)
    special_family_types = {}
    for s in all_students:
        if s.special_family:
            for t in s.special_family.split(','):
                t = t.strip()
                if t:
                    if t not in special_family_types:
                        special_family_types[t] = []
                    special_family_types[t].append(s.name)
    special_physical_count = sum(1 for s in all_students if s.special_physical == '是')
    special_physical_notes = [{'name': s.name, 'note': s.special_physical_note} for s in all_students if s.special_physical == '是' and s.special_physical_note]
    withdrawn_count = sum(1 for s in Student.query.filter_by(semester_id=sem_id) if s.status == 'withdrawn')
    withdrawn_students = [s for s in all_students if s.status == 'withdrawn']

    # 总人数排除流失
    active_students = [s for s in all_students if s.status != 'withdrawn']

    # 年龄分布
    ages = [get_age(s.id_card) for s in all_students if get_age(s.id_card) is not None]
    age_ranges = {'15岁以下': 0, '16-17岁': 0, '18-19岁': 0, '20岁以上': 0}
    for a in ages:
        if a < 16:
            age_ranges['15岁以下'] += 1
        elif a in (16, 17):
            age_ranges['16-17岁'] += 1
        elif a in (18, 19):
            age_ranges['18-19岁'] += 1
        else:
            age_ranges['20岁以上'] += 1

    avg_age = round(sum(ages) / len(ages), 1) if ages else 0
    max_age = max(ages) if ages else 0
    min_age = min(ages) if ages else 0

    # 户籍地区域
    region_counts = {}
    for s in all_students:
        r = get_region_from_idcard(s.id_card)
        if r:
            region_counts[r] = region_counts.get(r, 0) + 1
    regions = sorted(region_counts.items(), key=lambda x: -x[1])[:10]
    regions = [(name, name, count) for name, count in regions]

    stats = {
        'total': total,
        'display_count': display_count,
        'poverty_count': poverty_count,
        'boarding_count': boarding_count,
        'day_count': day_count,
        'minority_count': minority_count,
        'agri_count': agri_count,
        'non_agri_count': non_agri_count,
        'special_family_count': special_family_count,
        'special_family_types': special_family_types,
        'special_physical_count': special_physical_count,
        'special_physical_notes': special_physical_notes,
        'withdrawn_count': withdrawn_count,
        'withdrawn_students': withdrawn_students,
        'active_count': len(active_students),
        'age_ranges': age_ranges,
        'avg_age': avg_age,
        'max_age': max_age,
        'min_age': min_age,
        'regions': regions,
    }

    return render_template('students.html', students=students, stats=stats,
                          filter_type=filter_type, filter_value=filter_value)


@app.route('/students/add', methods=['GET', 'POST'])
def student_add():
    if request.method == 'POST':
        student = Student(
            name=request.form.get('name', '').strip(),
            student_id=generate_student_id(),
            class_name=request.form.get('class_name', ''),
            gender=request.form.get('gender', ''),
            id_card=request.form.get('id_card', ''),
            phone=request.form.get('phone', ''),
            parent_name=request.form.get('parent_name', ''),
            parent_phone=request.form.get('parent_phone', ''),
            parent_relation=request.form.get('parent_relation', ''),
            ethnic=request.form.get('ethnic', ''),
            political_status=request.form.get('political_status', ''),
            graduate_school=request.form.get('graduate_school', ''),
            residence_type=request.form.get('residence_type', ''),
            live_mode=request.form.get('live_mode', ''),
            poverty_status=request.form.get('poverty_status', ''),
            address=request.form.get('address', ''),
            major=request.form.get('major', ''),
            notes=request.form.get('notes', ''),
            dormitory=request.form.get('dormitory', ''),
            special_family=','.join(request.form.getlist('special_family')),
            special_family_note=request.form.get('special_family_note', ''),
            special_physical=request.form.get('special_physical', '否'),
            special_physical_note=request.form.get('special_physical_note', ''),
            remark=request.form.get('remark', ''),
            semester_id=get_current_semester_id(),
        )
        db.session.add(student)
        db.session.commit()
        flash(f'学生 {student.name} 添加成功，学号 {student.student_id}')
        return redirect(url_for('student_list'))
    return render_template('student_form.html', student=None)


@app.route('/students/<int:id>/edit', methods=['GET', 'POST'])
def student_edit(id):
    student = Student.query.get_or_404(id)
    if request.method == 'POST':
        student.name = request.form.get('name', '').strip()
        student.class_name = request.form.get('class_name', '')
        student.gender = request.form.get('gender', '')
        student.id_card = request.form.get('id_card', '')
        student.phone = request.form.get('phone', '')
        student.parent_name = request.form.get('parent_name', '')
        student.parent_phone = request.form.get('parent_phone', '')
        student.parent_relation = request.form.get('parent_relation', '')
        student.ethnic = request.form.get('ethnic', '')
        student.political_status = request.form.get('political_status', '')
        student.graduate_school = request.form.get('graduate_school', '')
        student.residence_type = request.form.get('residence_type', '')
        student.live_mode = request.form.get('live_mode', '')
        student.poverty_status = request.form.get('poverty_status', '')
        student.address = request.form.get('address', '')
        student.major = request.form.get('major', '')
        student.notes = request.form.get('notes', '')
        student.dormitory = request.form.get('dormitory', '')
        student.special_family = ','.join(request.form.getlist('special_family'))
        student.special_family_note = request.form.get('special_family_note', '')
        student.special_physical = request.form.get('special_physical', '否')
        student.special_physical_note = request.form.get('special_physical_note', '')
        student.remark = request.form.get('remark', '')
        db.session.commit()
        flash(f'学生 {student.name} 信息已更新')
        return redirect(url_for('student_list'))
    return render_template('student_form.html', student=student)


@app.route('/students/<int:id>/delete')
def student_delete(id):
    student = Student.query.get_or_404(id)
    name = student.name
    # 删除关联记录
    Attendance.query.filter_by(student_id=id).delete()
    Grade.query.filter_by(student_id=id).delete()
    TrainingRecord.query.filter_by(student_id=id).delete()
    db.session.delete(student)
    db.session.commit()
    flash(f'学生 {name} 已删除')
    return redirect(url_for('student_list'))


@app.route('/students/import', methods=['POST'])
def student_import():
    file = request.files.get('file')
    if not file:
        flash('请选择文件')
        return redirect(url_for('student_list'))

    filename = file.filename
    if not (filename.endswith('.xls') or filename.endswith('.xlsx')):
        flash('仅支持 .xls 和 .xlsx 格式')
        return redirect(url_for('student_list'))

    temp_path = os.path.join(UPLOAD_FOLDER, f'_import_{random.randint(1000,9999)}_{filename}')
    file.save(temp_path)

    try:
        imported = 0
        if temp_path.endswith('.xls'):
            import xlrd
            wb = xlrd.open_workbook(temp_path)
            sheet = wb.sheet_by_index(0)
            for r in range(sheet.nrows):
                if r < 2:
                    continue
                row = [str(sheet.cell_value(r, c)).strip() for c in range(min(sheet.ncols, 31))]
                if not row or len(row) < 2 or not row[1]:
                    continue
                name = row[1]
                if not name or len(name) > 20:
                    continue
                stu_id = str(row[30]).strip().split('.')[0] if len(row) > 30 and row[30] else str(row[3]).strip().upper() if row[3] else ''
                dup = Student.query.filter_by(name=name, id_card=stu_id, semester_id=get_current_semester_id()).first()
                if dup:
                    continue
                student = Student(
                    name=name, student_id=generate_student_id(),
                    class_name=row[11] if len(row) > 11 else '',
                    gender=row[2] if len(row) > 2 else '',
                    id_card=stu_id,
                    parent_name=row[12] if len(row) > 12 else '',
                    parent_phone=str(row[15]).split('.')[0] if len(row) > 15 and row[15] else '',
                    parent_relation=row[14] if len(row) > 14 else '',
                    graduate_school=row[16] if len(row) > 16 else '',
                    live_mode=row[17] if len(row) > 17 else '',
                    poverty_status=row[18] if len(row) > 18 else '',
                    residence_type=row[5] if len(row) > 5 else '',
                    address=row[25] if len(row) > 25 else '',
                    ethnic=row[4] if len(row) > 4 else '',
                    semester_id=get_current_semester_id(),
                )
                db.session.add(student)
                imported += 1
            db.session.commit()
            flash(f'成功导入 {imported} 名学生')
        else:
            from openpyxl import load_workbook
            wb = load_workbook(temp_path, read_only=True, data_only=True)
            sheet = wb.active
            for r, row in enumerate(sheet.iter_rows(values_only=True)):
                if r < 2:
                    continue
                if not row or len(row) < 2 or not row[1]:
                    continue
                name = str(row[1]).strip()
                if not name or len(name) > 20:
                    continue
                stu_id = ''
                if len(row) > 30 and row[30]:
                    stu_id = str(row[30]).strip().split('.')[0]
                elif len(row) > 3 and row[3]:
                    stu_id = str(row[3]).strip().upper()
                if Student.query.filter_by(name=name, id_card=stu_id, semester_id=get_current_semester_id()).first():
                    continue
                student = Student(
                    name=name, student_id=generate_student_id(),
                    class_name=str(row[11]).strip() if len(row) > 11 and row[11] else '',
                    gender=str(row[2]).strip() if len(row) > 2 and row[2] else '',
                    id_card=stu_id,
                    parent_name=str(row[12]).strip() if len(row) > 12 and row[12] else '',
                    parent_phone=str(row[15]).strip().split('.')[0] if len(row) > 15 and row[15] else '',
                    parent_relation=str(row[14]).strip() if len(row) > 14 and row[14] else '',
                    graduate_school=str(row[16]).strip() if len(row) > 16 and row[16] else '',
                    live_mode=str(row[17]).strip() if len(row) > 17 and row[17] else '',
                    poverty_status=str(row[18]).strip() if len(row) > 18 and row[18] else '',
                    residence_type=str(row[5]).strip() if len(row) > 5 and row[5] else '',
                    address=str(row[25]).strip() if len(row) > 25 and row[25] else '',
                    ethnic=str(row[4]).strip() if len(row) > 4 and row[4] else '',
                    semester_id=get_current_semester_id(),
                )
                db.session.add(student)
                imported += 1
            wb.close()
            db.session.commit()
            flash(f'成功导入 {imported} 名学生')
    except Exception as e:
        flash(f'导入失败: {e}')
    finally:
        try:
            os.remove(temp_path)
        except:
            pass
    return redirect(url_for('student_list'))

# ── 处分管理 ──
# ══════════════════════════════════════════════

def discipline_save_image(file, student_id, tag):
    """保存处分相关图片"""
    if not file or not file.filename:
        return ''
    ext = file.filename.rsplit('.', 1)[-1].lower() if '.' in file.filename else 'jpg'
    if ext not in ('jpg', 'jpeg', 'png', 'gif', 'webp'):
        return ''
    name = f'discipline_{student_id}_{tag}_{datetime.now().strftime("%Y%m%d%H%M%S")}_{random.randint(1000,9999)}.{ext}'
    path = os.path.join(UPLOAD_FOLDER, name)
    file.save(path)
    return f'uploads/{name}'

@app.route('/discipline')
def discipline_list():
    sem_id = get_current_semester_id()
    q = Discipline.query
    if sem_id:
        q = q.filter_by(semester_id=sem_id)
    records = q.order_by(Discipline.incident_date.desc()).all()
    students = get_semester_students()
    return render_template('discipline.html', records=records, students=students, now=datetime.now())


@app.route('/discipline/add', methods=['POST'])
def discipline_add():
    student_id = request.form.get('student_id', type=int)
    if not student_id or not Student.query.get(student_id):
        flash('请选择学生')
        return redirect(url_for('discipline_list'))
    try:
        incident_date = date.fromisoformat(request.form['incident_date'])
    except:
        incident_date = date.today()
    
    # 处理上传图片（支持多图）
    notif_files = request.files.getlist('image_notification')
    letter_files = request.files.getlist('image_letter')
    notif_paths = []
    for f in notif_files:
        p = discipline_save_image(f, student_id, 'notif')
        if p: notif_paths.append(p)
    letter_paths = []
    for f in letter_files:
        p = discipline_save_image(f, student_id, 'letter')
        if p: letter_paths.append(p)
    
    db.session.add(Discipline(
        student_id=student_id, incident_date=incident_date,
        location=request.form.get('location', ''),
        reason=request.form.get('reason', ''),
        punishment=request.form.get('punishment', ''),
        expiry_date=calc_expiry(incident_date),
        image_notification=json.dumps(notif_paths, ensure_ascii=False),
        image_letter=json.dumps(letter_paths, ensure_ascii=False),
        notes=request.form.get('notes', ''),
        semester_id=get_current_semester_id()
    ))
    db.session.commit()
    flash('处分记录已添加')
    return redirect(url_for('discipline_list'))


@app.route('/discipline/edit', methods=['POST'])
def discipline_edit_form():
    id = request.form.get('id', type=int)
    if not id:
        flash('参数错误')
        return redirect(url_for('discipline_list'))
    d = Discipline.query.get_or_404(id)
    try:
        d.incident_date = date.fromisoformat(request.form['incident_date'])
    except:
        pass
    d.location = request.form.get('location', '')
    d.reason = request.form.get('reason', '')
    d.punishment = request.form.get('punishment', '')
    d.notes = request.form.get('notes', '')
    notif_files = request.files.getlist('image_notification')
    letter_files = request.files.getlist('image_letter')
    new_notif = [discipline_save_image(f, d.student_id, 'notif') for f in notif_files if f.filename]
    new_letter = [discipline_save_image(f, d.student_id, 'letter') for f in letter_files if f.filename]
    if new_notif:
        existing = json.loads(d.image_notification) if d.image_notification else []
        d.image_notification = json.dumps(existing + new_notif, ensure_ascii=False)
    if new_letter:
        existing = json.loads(d.image_letter) if d.image_letter else []
        d.image_letter = json.dumps(existing + new_letter, ensure_ascii=False)
    expiry = calc_expiry(d.incident_date)
    if expiry:
        d.expiry_date = expiry
    db.session.commit()
    flash('处分记录已更新')
    return redirect(url_for('discipline_list'))


@app.route('/discipline/<int:id>/delete')
def discipline_delete(id):
    d = Discipline.query.get_or_404(id)
    db.session.delete(d)
    db.session.commit()
    flash('处分记录已删除')
    return redirect(url_for('discipline_list'))


@app.route('/export/discipline')
def export_discipline():
    fmt = request.args.get('format', 'excel')
    filter_student = request.args.get('student', '').strip()
    filter_start = request.args.get('start', '')
    filter_end = request.args.get('end', '')
    
    query = Discipline.query
    if filter_start:
        try: query = query.filter(Discipline.incident_date >= date.fromisoformat(filter_start))
        except: pass
    if filter_end:
        try: query = query.filter(Discipline.incident_date <= date.fromisoformat(filter_end))
        except: pass
    if filter_student:
        qs = Student.query.filter(Student.name.contains(filter_student))
        qs = qs.filter_by(semester_id=get_current_semester_id())
        students = qs.all()
        if students:
            query = query.filter(Discipline.student_id.in_([s.id for s in students]))
    
    records = query.order_by(Discipline.incident_date.desc()).all()
    
    if fmt == 'pdf':
        return _export_discipline_pdf(records)
    else:
        return _export_discipline_excel(records)


def _export_discipline_pdf(records):
    """处分PDF导出（含图片，每页最多2张）"""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer, Image as RLImage,
                                    PageBreak, Table, TableStyle)
    from reportlab.lib import colors
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    
    chinese_font = 'Helvetica'
    for fp in ['/System/Library/Fonts/PingFang.ttc', '/System/Library/Fonts/STHeiti Light.ttc']:
        if os.path.exists(fp):
            try:
                pdfmetrics.registerFont(TTFont('CnFont', fp)); chinese_font = 'CnFont'; break
            except: continue
    
    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=A4,
                           topMargin=15*mm, bottomMargin=15*mm,
                           leftMargin=15*mm, rightMargin=15*mm)
    style = ParagraphStyle('N', fontName=chinese_font, fontSize=10, leading=14)
    
    elements = []
    for r in records:
        elements.append(Paragraph(
            f'<b>{r.student.name}</b> {r.incident_date} | {r.location or "-"}<br/>'
            f'事由：{r.reason}<br/>处分：{r.punishment} | 到期：{r.expiry_date or "无"}',
            style))
        elements.append(Spacer(1, 6))
        
        # 显示通知照片
        if r.image_notification:
            imgs = json.loads(r.image_notification)
            for img_path in imgs[:2]:  # 每页最多2张
                full = os.path.join(BASE_DIR, 'static', img_path)
                if os.path.exists(full):
                    try:
                        elements.append(RLImage(full, width=400, height=300, kind='proportional'))
                        elements.append(Spacer(1, 4))
                    except:
                        elements.append(Paragraph(f'[图片]', style))
        # 显示告知书照片
        if r.image_letter:
            imgs = json.loads(r.image_letter)
            for img_path in imgs[:2]:
                full = os.path.join(BASE_DIR, 'static', img_path)
                if os.path.exists(full):
                    try:
                        elements.append(RLImage(full, width=400, height=300, kind='proportional'))
                        elements.append(Spacer(1, 4))
                    except:
                        elements.append(Paragraph(f'[图片]', style))
        elements.append(PageBreak())
    
    doc.build(elements)
    output.seek(0)
    return send_file(output, mimetype='application/pdf', as_attachment=True,
                    download_name=f'处分记录_{date.today().isoformat()}.pdf')


def _export_discipline_excel(records):
    """处分Excel导出（含嵌入图片）"""
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.styles import Font, Alignment, Border, Side
    
    wb = Workbook()
    ws = wb.active
    ws.title = '处分记录'
    
    headers = ['姓名', '性别', '日期', '地点', '事由', '处分', '到期日', '备注']
    hf = Font(bold=True, size=11)
    thin = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
    align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = hf; c.border = thin; c.alignment = align
    
    for ri, r in enumerate(records, 2):
        ws.cell(row=ri, column=1, value=r.student.name).border = thin
        ws.cell(row=ri, column=2, value=r.student.gender or '').border = thin
        ws.cell(row=ri, column=3, value=str(r.incident_date)).border = thin
        ws.cell(row=ri, column=4, value=r.location or '').border = thin
        ws.cell(row=ri, column=5, value=r.reason).border = thin
        ws.cell(row=ri, column=6, value=r.punishment).border = thin
        ws.cell(row=ri, column=7, value=str(r.expiry_date or '')).border = thin
        ws.cell(row=ri, column=8, value=r.notes or '').border = thin
        
        row_imgs = []
        if r.image_notification:
            row_imgs.extend(json.loads(r.image_notification))
        if r.image_letter:
            row_imgs.extend(json.loads(r.image_letter))
        
        col = 9
        for img_path in row_imgs:
            full = os.path.join(BASE_DIR, 'static', img_path)
            if os.path.exists(full):
                try:
                    img = XLImage(full)
                    aspect = img.width / img.height if img.height else 1
                    img.width = min(300, img.width)
                    img.height = img.width / aspect
                    ws.add_image(img, f'{chr(64+col)}{ri}')
                    ws.column_dimensions[chr(64+col)].width = max(20, img.width * 0.15)
                    col += 1
                except:
                    pass
        ws.row_dimensions[ri].height = max(80, 150)
    
    for ci, w in enumerate([12, 8, 12, 14, 30, 12, 12, 20], 1):
        ws.column_dimensions[chr(64 + ci)].width = w
    
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf, as_attachment=True, download_name=f'处分记录_{date.today().isoformat()}.xlsx')


# ── 考勤管理 ──

@app.route('/attendance', methods=['GET', 'POST'])
def attendance():
    students = [s for s in get_semester_students() if s.status != 'withdrawn']
    today = date.today()
    query_date_str = request.args.get('date', today.isoformat())

    try:
        query_date = date.fromisoformat(query_date_str)
    except ValueError:
        query_date = today

    if request.method == 'POST':
        form_date_str = request.form.get('date', query_date_str)
        try:
            form_date = date.fromisoformat(form_date_str)
        except ValueError:
            form_date = today

        for s in students:
            status = request.form.get(f'status_{s.id}', 'present')
            reason = request.form.get(f'reason_{s.id}', '')
            image_file = request.files.get(f'image_{s.id}')
            
            row_period = int(request.form.get(f'period_sum_{s.id}', 2047))
            row_period = max(0, min(2047, row_period))

            # 每人每天一条记录，按日期查找
            existing = Attendance.query.filter_by(
                student_id=s.id, date=form_date
            ).first()

            if existing:
                existing.status = status
                existing.reason = reason
                existing.period = row_period
                if image_file and image_file.filename:
                    ext = image_file.filename.rsplit('.', 1)[-1].lower() if '.' in image_file.filename else 'jpg'
                    safe_name = f'att_{s.id}_{form_date_str.replace("-","")}_{row_period}_{random.randint(1000,9999)}.{ext}'
                    save_path = os.path.join(UPLOAD_FOLDER, safe_name)
                    image_file.save(save_path)
                    existing.image_path = f'uploads/{safe_name}'
            else:
                image_path = ''
                if image_file and image_file.filename:
                    ext = image_file.filename.rsplit('.', 1)[-1].lower() if '.' in image_file.filename else 'jpg'
                    safe_name = f'att_{s.id}_{form_date_str.replace("-","")}_{row_period}_{random.randint(1000,9999)}.{ext}'
                    save_path = os.path.join(UPLOAD_FOLDER, safe_name)
                    image_file.save(save_path)
                    image_path = f'uploads/{safe_name}'
                db.session.add(Attendance(
                    student_id=s.id, date=form_date, status=status,
                    reason=reason, period=row_period,
                    image_path=image_path,
                    semester_id=get_current_semester_id()
                ))
        db.session.commit()
        flash(f'{form_date_str} 考勤已保存')
        return redirect(url_for('attendance', date=form_date_str))

    # GET: 加载该日期的所有记录
    sem_id = get_current_semester_id()
    records = Attendance.query.filter_by(date=query_date).all()
    if sem_id:
        records = [r for r in records if r.semester_id == sem_id]
    existing_map = {r.student_id: r for r in records}
    return render_template('attendance.html', students=students, records=records,
                          existing_map=existing_map, query_date=query_date_str)


@app.route('/attendance/stats')
def attendance_stats():
    sem_id = get_current_semester_id()
    students = [s for s in get_semester_students() if s.status != 'withdrawn']
    stats_data = []
    for s in students:
        q = Attendance.query.filter_by(student_id=s.id)
        if sem_id:
            q = q.filter_by(semester_id=sem_id)
        records = q.all()
        sick_hours = sum(popcount(r.period) for r in records if r.status == 'sick')
        personal_hours = sum(popcount(r.period) for r in records if r.status == 'personal')
        truant_hours = sum(popcount(r.period) for r in records if r.status == 'truant')
        late_count = sum(1 for r in records if r.status == 'late')
        total_absent = sick_hours + personal_hours + truant_hours
        stats_data.append({
            'student': s,
            'sick_hours': sick_hours,
            'personal_hours': personal_hours,
            'truant_hours': truant_hours,
            'late_count': late_count,
            'total_absent': total_absent,
        })
    
    # 最近15天每日缺勤统计（用于图表）
    today = date.today()
    chart_dates = []
    chart_truant = []
    chart_sick = []
    chart_personal = []
    for i in range(14, -1, -1):
        d = today - timedelta(days=i)
        chart_dates.append(d.strftime('%m-%d'))
        day_records = Attendance.query.filter_by(date=d).all()
        if sem_id:
            day_records = [r for r in day_records if r.semester_id == sem_id]
        chart_truant.append(sum(popcount(r.period) for r in day_records if r.status == 'truant'))
        chart_sick.append(sum(popcount(r.period) for r in day_records if r.status == 'sick'))
        chart_personal.append(sum(popcount(r.period) for r in day_records if r.status == 'personal'))
    
    total_students = len([s for s in get_semester_students() if s.status != 'withdrawn'])
    
    return render_template('attendance_stats.html',
                          students=stats_data,
                          warn_hours=TRUANCY_WARN_HOURS,
                          chart_dates=chart_dates, chart_truant=chart_truant,
                          chart_sick=chart_sick, chart_personal=chart_personal,
                          total_students=total_students)


@app.route('/attendance/weekly')
def attendance_weekly():
    today = date.today()
    week_offset = request.args.get('offset', 0, type=int)
    monday = today - timedelta(days=today.weekday()) + timedelta(weeks=week_offset)
    sunday = monday + timedelta(days=6)
    sem_id = get_current_semester_id()
    
    week_dates = [monday + timedelta(days=i) for i in range(7)]
    week_names = ['周一', '周二', '周三', '周四', '周五', '周六', '周日']
    
    students = [s for s in get_semester_students() if s.status != 'withdrawn']
    week_data = []
    for s in students:
        day_records = {}
        for d in week_dates:
            r = Attendance.query.filter_by(student_id=s.id, date=d).first()
            if r and sem_id and r.semester_id != sem_id:
                r = None
            day_records[d.isoformat()] = r
        week_data.append({'student': s, 'records': day_records})
    
    return render_template('attendance_weekly.html', students=week_data,
                          week_dates=week_dates, week_names=week_names,
                          monday=monday, sunday=sunday, week_offset=week_offset)

# ══════════════════════════════════════════════
# 违纪记录
# ══════════════════════════════════════════════

@app.route('/violation')
@login_required
def violation_list():
    sem_id = get_current_semester_id()
    q = ViolationRecord.query
    if sem_id: q = q.filter_by(semester_id=sem_id)
    records = q.order_by(ViolationRecord.date.desc()).all()
    students = Student.query.order_by(Student.name).all()
    if sem_id: students = [s for s in students if s.semester_id == sem_id]
    days_dist = {}
    for r in records: days_dist[r.reflection_days] = days_dist.get(r.reflection_days, 0) + 1
    return render_template('violation.html', records=records, students=students, stats={'total':len(records),'days_dist':days_dist})

@app.route('/violation/add', methods=['POST'])
@login_required
def violation_add():
    sid = request.form.get('student_id', type=int)
    d = date.today()
    try: d = datetime.strptime(request.form.get('date',''), '%Y-%m-%d').date()
    except: pass
    rd = min(15, max(1, request.form.get('reflection_days', type=int, default=1)))
    v = ViolationRecord(student_id=sid, date=d, reason=request.form.get('reason',''),
                        reflection_days=rd, reflection_start=d, reflection_end=add_date(d, rd),
                        notes=request.form.get('notes',''), semester_id=get_current_semester_id())
    db.session.add(v); db.session.commit()
    flash('违纪记录已添加'); return redirect(url_for('violation_list'))

@app.route('/violation/<int:id>/edit', methods=['GET','POST'])
@login_required
def violation_edit(id):
    v = ViolationRecord.query.get_or_404(id)
    if request.method == 'POST':
        v.student_id = request.form.get('student_id', type=int)
        try: v.date = datetime.strptime(request.form.get('date',''), '%Y-%m-%d').date()
        except: pass
        v.reason = request.form.get('reason','')
        v.reflection_days = min(15, max(1, request.form.get('reflection_days', type=int, default=1)))
        v.reflection_start = v.date; v.reflection_end = add_date(v.date, v.reflection_days)
        v.notes = request.form.get('notes','')
        db.session.commit(); flash('已更新'); return redirect(url_for('violation_list'))
    students = Student.query.order_by(Student.name).all()
    return render_template('violation_edit.html', v=v, students=students)

@app.route('/violation/<int:id>/delete')
@login_required
def violation_delete(id):
    v = ViolationRecord.query.get_or_404(id); db.session.delete(v); db.session.commit()
    flash('已删除'); return redirect(url_for('violation_list'))


@app.route('/export/special-family')
@login_required
def export_special_family():
    from openpyxl import Workbook, styles; import io
    sem_id=get_current_semester_id()
    all_students=Student.query.order_by(Student.name).all()
    if sem_id: all_students=[s for s in all_students if s.semester_id==sem_id]
    students=[s for s in all_students if s.special_family]
    wb=Workbook();ws=wb.active;ws.title='特殊家庭台账'
    thin=styles.Side(style='thin');bd=styles.Border(left=thin,right=thin,top=thin,bottom=thin)
    hf=styles.Font(name='黑体',size=10,bold=True);df=styles.Font(name='宋体',size=10)
    ca=styles.Alignment(horizontal='center',vertical='center',wrap_text=True)
    ws.merge_cells('A1:S1')
    ws['A1']='榆中县特殊家庭学生信息台账';ws['A1'].font=styles.Font(name='黑体',size=14,bold=True);ws['A1'].alignment=ca
    ws.merge_cells('A2:S2')
    ws['A2']='学校（盖章）：榆中县职业教育中心';ws['A2'].font=styles.Font(name='宋体',size=10);ws['A2'].alignment=ca
    headers=['序号','学生姓名','性别','民族','出生年月','所在学校','班级','家庭详细地址','家长姓名','联系电话','单亲','留守','离异','孤儿','残疾','贫困','其他','特殊情况说明','备注']
    for i,h in enumerate(headers,1):
        c=ws.cell(row=3,column=i,value=h);c.font=hf;c.border=bd;c.alignment=ca
    for idx,s in enumerate(students,1):
        r=idx+3;sf=s.special_family.split(',') if s.special_family else []
        vals=[idx,s.name,s.gender,s.ethnic,'','榆中县职业教育中心',s.class_name,s.address,s.parent_name,s.parent_phone]
        for ft in ['单亲','留守','离异','孤儿','残疾','贫困','其他']: vals.append('√' if ft in sf else '')
        vals.append(s.special_family_note or '');vals.append(s.remark or '')
        if s.id_card and len(s.id_card)>=10: vals[4]=s.id_card[6:10]+'-'+s.id_card[10:12]+'-'+s.id_card[12:14]
        for ci,v in enumerate(vals,1):
            c=ws.cell(row=r,column=ci,value=v);c.border=bd;c.font=df;c.alignment=ca
    for i,w in enumerate([6,8,6,6,12,18,12,30,8,14,5,5,5,5,5,5,5,20,20],1):
        ws.column_dimensions[chr(64+i) if i<=26 else 'A'].width=w
    buf=io.BytesIO();wb.save(buf);buf.seek(0)
    return send_file(buf,as_attachment=True,download_name='特殊家庭学生信息台账.xlsx',mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/export/dormitory')
@login_required
def export_dormitory():
    from openpyxl import Workbook, styles; import io
    sem_id=get_current_semester_id()
    students=Student.query.order_by(Student.dormitory,Student.name).all()
    if sem_id: students=[s for s in students if s.semester_id==sem_id and s.dormitory]
    wb=Workbook();ws=wb.active;ws.title='宿舍名单'
    thin=styles.Side(style='thin');bd=styles.Border(left=thin,right=thin,top=thin,bottom=thin)
    hf=styles.Font(name='黑体',size=11,bold=True);df=styles.Font(name='宋体',size=10)
    ca=styles.Alignment(horizontal='center',vertical='center',wrap_text=True)
    ws.merge_cells('A1:D1')
    ws['A1']='学生宿舍名单';ws['A1'].font=styles.Font(name='黑体',size=14,bold=True);ws['A1'].alignment=ca
    for i,h in enumerate(['宿舍号','姓名','班级','性别'],1):
        ws.cell(row=2,column=i,value=h).font=hf;ws.cell(row=2,column=i).alignment=ca;ws.cell(row=2,column=i).border=bd
    for idx,s in enumerate(students,3):
        ws.cell(row=idx,column=1,value=s.dormitory).border=bd;ws.cell(row=idx,column=1).font=df
        ws.cell(row=idx,column=2,value=s.name).border=bd;ws.cell(row=idx,column=2).font=df
        ws.cell(row=idx,column=3,value=s.class_name).border=bd;ws.cell(row=idx,column=3).font=df
        ws.cell(row=idx,column=4,value=s.gender).border=bd;ws.cell(row=idx,column=4).font=df
        for c in range(1,5): ws.cell(row=idx,column=c).alignment=ca
    ws.column_dimensions['A'].width=12;ws.column_dimensions['B'].width=10
    ws.column_dimensions['C'].width=22;ws.column_dimensions['D'].width=8
    buf=io.BytesIO();wb.save(buf);buf.seek(0)
    return send_file(buf,as_attachment=True,download_name='宿舍名单.xlsx',mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/export/violations')
@login_required
def export_violations():
    from openpyxl import Workbook, styles; import io
    sem_id = get_current_semester_id()
    q = ViolationRecord.query
    if sem_id: q = q.filter_by(semester_id=sem_id)
    records = q.order_by(ViolationRecord.date.desc()).all()
    wb = Workbook(); ws = wb.active; ws.title = '违纪记录'
    thin = styles.Side(style='thin'); bd = styles.Border(left=thin,right=thin,top=thin,bottom=thin)
    hf = styles.Font(name='黑体', size=10, bold=True); df = styles.Font(name='宋体', size=10)
    ca = styles.Alignment(horizontal='center', vertical='center', wrap_text=True)
    for i,h in enumerate(['序号','姓名','日期','违纪事由','反省天数','反省起止','备注'],1):
        c=ws.cell(row=1,column=i,value=h);c.font=hf;c.border=bd;c.alignment=ca
    for idx,r in enumerate(records,1):
        row=idx+1
        vals=[idx,r.student.name,r.date.isoformat(),r.reason,f'{r.reflection_days}天',f'{r.reflection_start}~{r.reflection_end}',r.notes]
        for ci,v in enumerate(vals,1):
            c=ws.cell(row=row,column=ci,value=v);c.border=bd;c.font=df;c.alignment=ca
    for i,w in enumerate([6,10,12,30,10,28,20],1): ws.column_dimensions[chr(64+i) if i<=26 else 'A'].width=w
    buf=io.BytesIO();wb.save(buf);buf.seek(0)
    return send_file(buf,as_attachment=True,download_name='违纪记录.xlsx',mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/export/violation-stats')
@login_required
def export_violation_stats():
    from openpyxl import Workbook, styles; import io
    sem_id = get_current_semester_id()
    q = ViolationRecord.query
    if sem_id: q = q.filter_by(semester_id=sem_id)
    records = q.all()
    wb = Workbook(); ws = wb.active; ws.title = '违纪统计'
    thin = styles.Side(style='thin'); bd = styles.Border(left=thin,right=thin,top=thin,bottom=thin)
    hf = styles.Font(name='黑体', size=10, bold=True); df = styles.Font(name='宋体', size=10)
    ca = styles.Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.merge_cells('A1:C1'); ws['A1']='违纪统计汇总'
    ws['A1'].font=styles.Font(name='黑体', size=14, bold=True); ws['A1'].alignment=ca
    for i,h in enumerate(['反省天数','记录数','占比'],1):
        c=ws.cell(row=2,column=i,value=h);c.font=hf;c.border=bd;c.alignment=ca
    total=len(records); dist={}
    for r in records: dist[r.reflection_days]=dist.get(r.reflection_days,0)+1
    row=3
    for d in sorted(dist):
        ws.cell(row=row,column=1,value=f'{d}天').border=bd;ws.cell(row=row,column=1).font=df
        ws.cell(row=row,column=1).alignment=ca
        ws.cell(row=row,column=2,value=dist[d]).border=bd;ws.cell(row=row,column=2).font=df
        ws.cell(row=row,column=2).alignment=ca
        ws.cell(row=row,column=3,value=f'{dist[d]/total*100:.1f}%' if total else '0%').border=bd
        ws.cell(row=row,column=3).font=df;ws.cell(row=row,column=3).alignment=ca
        row+=1
    ws.column_dimensions['A'].width=12;ws.column_dimensions['B'].width=10;ws.column_dimensions['C'].width=10
    buf=io.BytesIO();wb.save(buf);buf.seek(0)
    return send_file(buf,as_attachment=True,download_name='违纪统计.xlsx',mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ══════════════════════════════════════════════
# 班费管理
# ══════════════════════════════════════════════

@app.route('/fund')
@login_required
def fund_list():
    sem_id = get_current_semester_id()
    q = ClassFund.query
    if sem_id: q = q.filter_by(semester_id=sem_id)
    records = q.order_by(ClassFund.date.desc()).all()
    income = sum(r.amount for r in records if r.type == 'income')
    expense = sum(r.amount for r in records if r.type == 'expense')
    return render_template('fund.html', records=records, income=income, expense=expense, balance=income-expense)

@app.route('/fund/add', methods=['POST'])
@login_required
def fund_add():
    d = date.today()
    try: d = datetime.strptime(request.form.get('date',''), '%Y-%m-%d').date()
    except: pass
    ft = request.form.get('type', 'income')
    amount = request.form.get('amount', type=float, default=0)
    f = ClassFund(date=d, type=ft, amount=abs(amount), reason=request.form.get('reason',''),
                  notes=request.form.get('notes',''), semester_id=get_current_semester_id())
    # 处理凭据图片
    if 'voucher' in request.files:
        file = request.files['voucher']
        if file and file.filename:
            ext = os.path.splitext(file.filename)[1] or '.jpg'
            filename = f'fund_{int(__import__("time").time())}{ext}'
            file.save(os.path.join(UPLOAD_FOLDER, filename))
            f.voucher = filename
    db.session.add(f); db.session.commit()
    flash(f'{"收入" if ft=="income" else "支出"} {amount:.2f} 元已记录'); return redirect(url_for('fund_list'))

@app.route('/fund/<int:id>/delete')
@login_required
def fund_delete(id):
    f = ClassFund.query.get_or_404(id)
    if f.voucher:
        try: os.remove(os.path.join(UPLOAD_FOLDER, f.voucher))
        except: pass
    db.session.delete(f); db.session.commit(); flash('已删除'); return redirect(url_for('fund_list'))

@app.route('/export/fund-xlsx')
@login_required
def export_fund_xlsx():
    """导出班费明细Excel"""
    from openpyxl import Workbook, styles
    from PIL import Image as PILImage
    import io
    sem_id = get_current_semester_id()
    q = ClassFund.query
    if sem_id: q = q.filter_by(semester_id=sem_id)
    records = q.order_by(ClassFund.date.desc()).all()
    
    wb = Workbook(); ws = wb.active; ws.title = '班费明细'
    thin = styles.Side(style='thin'); bd = styles.Border(left=thin,right=thin,top=thin,bottom=thin)
    hf = styles.Font(name='黑体', size=10, bold=True); df = styles.Font(name='宋体', size=10)
    ca = styles.Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    for i, h in enumerate(['日期','类型','金额','事由','凭据','备注'], 1):
        c = ws.cell(row=1, column=i, value=h); c.font = hf; c.border = bd; c.alignment = ca
    
    income = sum(r.amount for r in records if r.type == 'income')
    expense = sum(r.amount for r in records if r.type == 'expense')
    ws.cell(row=2, column=1, value='总收入').font = hf; ws.cell(row=2, column=1).border = bd
    ws.merge_cells('A2:B2')
    ws.cell(row=2, column=3, value=income).border = bd; ws.cell(row=2, column=3).font = df
    ws.cell(row=3, column=1, value='总支出').font = hf; ws.cell(row=3, column=1).border = bd
    ws.merge_cells('A3:B3')
    ws.cell(row=3, column=3, value=expense).border = bd; ws.cell(row=3, column=3).font = df
    ws.cell(row=4, column=1, value='结余').font = hf; ws.cell(row=4, column=1).border = bd
    ws.merge_cells('A4:B4')
    ws.cell(row=4, column=3, value=income-expense).border = bd; ws.cell(row=4, column=3).font = df
    
    row = 5
    ws.cell(row=row, column=1, value='--- 明细 ---').font = hf; ws.merge_cells(f'A{row}:F{row}')
    row += 1
    for r in records:
        ws.cell(row=row, column=1, value=r.date.isoformat()).border = bd; ws.cell(row=row, column=1).font = df
        ws.cell(row=row, column=1).alignment = ca
        ws.cell(row=row, column=2, value='收入' if r.type=='income' else '支出').border = bd
        ws.cell(row=row, column=2).font = df; ws.cell(row=row, column=2).alignment = ca
        ws.cell(row=row, column=3, value=r.amount).border = bd; ws.cell(row=row, column=3).font = df
        ws.cell(row=row, column=4, value=r.reason).border = bd; ws.cell(row=row, column=4).font = df
        if r.voucher:
            vpath = os.path.join(UPLOAD_FOLDER, r.voucher)
            if os.path.exists(vpath):
                from openpyxl.drawing.image import Image as XLImage
                try:
                    img = XLImage(vpath)
                    img.width = min(200, img.width)
                    img.height = img.width * PILImage.open(vpath).size[1] / PILImage.open(vpath).size[0]
                    ws.add_image(img, f'E{row}')
                except: pass
            ws.cell(row=row, column=5, value='有凭据').border = bd
        ws.cell(row=row, column=6, value=r.notes).border = bd; ws.cell(row=row, column=6).font = df
        row += 1
    
    ws.column_dimensions['A'].width = 12; ws.column_dimensions['B'].width = 8
    ws.column_dimensions['C'].width = 10; ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 12; ws.column_dimensions['F'].width = 20
    
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf, as_attachment=True, download_name='班费明细.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ══════════════════════════════════════════════
# 成绩管理（综合评价）
# ══════════════════════════════════════════════

@app.route('/grades')
def grades():
    sem_id = get_current_semester_id()
    students = get_semester_students()
    subjects = get_subjects()
    if not subjects:
        return render_template('grades.html', students=[], subjects=[],
                              current_subject='', avg=0)
    current_subject = request.args.get('subject', subjects[0])
    if current_subject not in subjects:
        current_subject = subjects[0]
    
    data = []
    total_comp = 0
    sub_count = 0
    for s in students:
        q = Grade.query.filter_by(student_id=s.id, subject=current_subject)
        if sem_id:
            q = q.filter_by(semester_id=sem_id)
        g = q.first()
        data.append({'student': s, 'grade': g})
        if g and g.comprehensive_score:
            total_comp += g.comprehensive_score
            sub_count += 1
    avg = round(total_comp / sub_count, 1) if sub_count > 0 else 0
    
    return render_template('grades.html', students=data, subjects=subjects,
                          current_subject=current_subject, avg=avg,
                          subject_objects=get_semester_subject_objects())


@app.route('/grades/save', methods=['POST'])
def grade_save():
    """保存当前科目的所有成绩"""
    subject = request.form.get('subject', '')
    if subject not in get_subjects():
        flash('科目无效')
        return redirect(url_for('grades'))
    
    for key, val in request.form.items():
        if key.startswith('perf_'):
            sid = int(key.split('_')[1])
            perf = float(val or 0)
            hw = float(request.form.get(f'hw_{sid}', 0) or 0)
            notes = float(request.form.get(f'notes_{sid}', 0) or 0)
            exam = float(request.form.get(f'exam_{sid}', 0) or 0)
            if perf == 0 and hw == 0 and notes == 0 and exam == 0:
                continue
            g = Grade.query.filter_by(student_id=sid, subject=subject).first()
            if not g:
                g = Grade(student_id=sid, subject=subject)
                db.session.add(g)
                g.semester_id = get_current_semester_id()
            g.performance_score = perf
            g.homework_score = hw
            g.notes_score = notes
            g.exam_score = exam
            g.calc_comprehensive()
    
    db.session.commit()
    flash(f'「{subject}」成绩已保存')
    return redirect(url_for('grades', subject=subject))


@app.route('/grades/import', methods=['POST'])
def grade_import():
    file = request.files.get('file')
    if not file:
        flash('请选择文件')
        return redirect(url_for('grades'))
    try:
        import openpyxl
        wb = openpyxl.load_workbook(file, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 5:
            flash('文件格式错误：数据行不足')
            return redirect(url_for('grades'))
        
        # 解析表头获取科目和列映射
        # R3: 科目名称行 (C3, C8, C13, ...)
        # R4: 评分类型行
        col_map = []  # [(col_idx, subject)]
        for c in range(2, min(len(rows[2] or []), 44)):
            cell = str(rows[2][c] or '').strip()
            if '科目：' in cell:
                sub = cell.replace('科目：', '').split('任课')[0].strip()
                col_map.append((c, sub))
        
        imported = 0
        for r in range(4, len(rows)):
            row = rows[r]
            if not row or not row[1]:
                continue
            name = str(row[1]).strip()
            student = Student.query.filter_by(name=name).first()
            if not student:
                continue
            
            overall = 0
            sub_count = 0
            for ci, sub in col_map:
                if ci + 3 < len(row):
                    perf = float(row[ci] or 0) if row[ci] else 0
                    hw = float(row[ci+1] or 0) if ci+1 < len(row) and row[ci+1] else 0
                    notes = float(row[ci+2] or 0) if ci+2 < len(row) and row[ci+2] else 0
                    exam = float(row[ci+3] or 0) if ci+3 < len(row) and row[ci+3] else 0
                    if perf == 0 and hw == 0 and notes == 0 and exam == 0:
                        continue
                    g = Grade.query.filter_by(student_id=student.id, subject=sub).first()
                    if not g:
                        g = Grade(student_id=student.id, subject=sub)
                        db.session.add(g)
                    g.performance_score = perf
                    g.homework_score = hw
                    g.notes_score = notes
                    g.exam_score = exam
                    g.calc_comprehensive()
                    overall += g.comprehensive_score
                    sub_count += 1
                    imported += 1
            
            if sub_count > 0:
                g_all = Grade.query.filter_by(student_id=student.id, subject='').first()
                if not g_all:
                    g_all = Grade(student_id=student.id, subject='')
                    db.session.add(g_all)
                g_all.overall_score = round(overall / sub_count, 1)
        
        db.session.commit()
        flash(f'成功导入 {imported} 条成绩记录')
    except Exception as e:
        flash(f'导入失败: {e}')
    return redirect(url_for('grades'))


@app.route('/grades/analysis')
def grade_analysis():
    sem_id = get_current_semester_id()
    subjects = get_subjects()
    averages = {}
    for sub in subjects:
        q = Grade.query.filter_by(subject=sub)
        if sem_id:
            q = q.filter_by(semester_id=sem_id)
        scores = [g.comprehensive_score for g in q.all() if g.comprehensive_score]
        averages[sub] = round(sum(scores) / len(scores), 1) if scores else 0
    gq = Grade.query.filter(Grade.overall_score > 0)
    if sem_id:
        gq = gq.filter_by(semester_id=sem_id)
    overs = [g.overall_score for g in gq.all()]
    avg_overall = round(sum(overs) / len(overs), 1) if overs else 0
    return render_template('grade_analysis.html', averages=averages, subjects=subjects, avg_overall=avg_overall)


# ══════════════════════════════════════════════
# 实训管理
# ══════════════════════════════════════════════

@app.route('/training')
def training():
    sem_id = get_current_semester_id()
    projects = get_semester_projects()
    students = get_semester_students()
    q = TrainingRecord.query
    if sem_id:
        q = q.filter_by(semester_id=sem_id)
    recent_records = q.order_by(TrainingRecord.created_at.desc()).limit(30).all()

    # latest_map: {student_id_project_id: record}
    latest_map = {}
    for s in students:
        for p in projects:
            rq = TrainingRecord.query.filter_by(student_id=s.id, project_id=p.id)
            if sem_id:
                rq = rq.filter_by(semester_id=sem_id)
            record = rq.order_by(TrainingRecord.created_at.desc()).first()
            if record:
                key = f'{s.id}_{p.id}'
                latest_map[key] = record

    gq = TrainingGroup.query
    if sem_id:
        gq = gq.filter_by(semester_id=sem_id)
    groups = gq.order_by(TrainingGroup.id).all()

    return render_template('training.html',
                          projects=projects,
                          students=students,
                          groups=groups,
                          recent_records=recent_records,
                          latest_map=latest_map,
                          now=datetime.now())


@app.route('/training/project/add', methods=['POST'])
def training_project_add():
    name = request.form.get('name', '').strip()
    if not name:
        flash('请输入项目名称')
        return redirect(url_for('training'))
    description = request.form.get('description', '')
    category = request.form.get('category', '')
    max_score = request.form.get('max_score', '100')
    try:
        max_score = float(max_score)
    except ValueError:
        max_score = 100

    project = TrainingProject(name=name, description=description,
                              max_score=max_score, category=category,
                              semester_id=get_current_semester_id())
    db.session.add(project)
    db.session.commit()
    flash(f'实训项目 {name} 已添加')
    return redirect(url_for('training'))


@app.route('/training/project/<int:id>/delete')
def training_project_delete(id):
    project = TrainingProject.query.get_or_404(id)
    TrainingRecord.query.filter_by(project_id=id).delete()
    name = project.name
    db.session.delete(project)
    db.session.commit()
    flash(f'实训项目 {name} 已删除')
    return redirect(url_for('training'))


@app.route('/training/record', methods=['POST'])
def training_record_add():
    group_id = request.form.get('group_id')
    student_id = request.form.get('student_id')
    project_id = request.form.get('project_id')
    score = request.form.get('score', '0')
    quality_notes = request.form.get('quality_notes', '')
    completion_date_str = request.form.get('completion_date', '')
    instructor_notes = request.form.get('instructor_notes', '')

    try:
        project_id = int(project_id)
        score = float(score)
    except (ValueError, TypeError):
        flash('参数错误')
        return redirect(url_for('training'))

    try:
        completion_date = date.fromisoformat(completion_date_str) if completion_date_str else None
    except ValueError:
        completion_date = None

    sem_id = get_current_semester_id()

    # 如果选择了分组，为组内所有学生创建实训记录
    if group_id:
        try:
            group_id = int(group_id)
        except (ValueError, TypeError):
            flash('参数错误')
            return redirect(url_for('training'))
        group = TrainingGroup.query.get(group_id)
        if not group:
            flash('分组不存在')
            return redirect(url_for('training'))
        students = group.students.all()
        if not students:
            flash('该分组没有成员')
            return redirect(url_for('training'))
        count = 0
        for s in students:
            record = TrainingRecord(
                student_id=s.id, project_id=project_id, score=score,
                quality_notes=quality_notes, completion_date=completion_date,
                instructor_notes=instructor_notes,
                group_id=group_id, semester_id=sem_id
            )
            db.session.add(record)
            count += 1
        db.session.commit()
        flash(f'已为分组 "{group.name}" 的 {count} 名学生添加实训记录')
    else:
        # 传统方式：为单个学生添加记录
        try:
            student_id = int(student_id)
        except (ValueError, TypeError):
            flash('请选择学生或分组')
            return redirect(url_for('training'))
        record = TrainingRecord(
            student_id=student_id, project_id=project_id, score=score,
            quality_notes=quality_notes, completion_date=completion_date,
            instructor_notes=instructor_notes,
            semester_id=sem_id
        )
        db.session.add(record)
        db.session.commit()
        flash('实训记录已添加')
    return redirect(url_for('training'))


# ══════════════════════════════════════════════
# 实训分组管理
# ══════════════════════════════════════════════

@app.route('/training/groups')
def training_groups():
    sem_id = get_current_semester_id()
    groups = TrainingGroup.query.filter_by(semester_id=sem_id).order_by(TrainingGroup.id).all()
    all_students = get_semester_students()
    # 已分配到任何分组的学生ID集合
    assigned_ids = set()
    for g in groups:
        for s in g.students:
            assigned_ids.add(s.id)
    # 未分配的学生（未被任何分组选中的学生）
    students = [s for s in all_students if s.id not in assigned_ids and s.status != 'withdrawn']
    return render_template('training_groups.html', groups=groups, students=students)


@app.route('/training/groups/add', methods=['POST'])
def training_group_add():
    name = request.form.get('name', '').strip()
    if not name:
        flash('请输入分组名称')
        return redirect(url_for('training_groups'))
    sem_id = get_current_semester_id()
    group = TrainingGroup(name=name, semester_id=sem_id)
    db.session.add(group)
    db.session.commit()
    flash(f'实训分组 "{name}" 已创建')
    return redirect(url_for('training_groups'))


@app.route('/training/groups/<int:id>/delete', methods=['POST'])
def training_group_delete(id):
    group = TrainingGroup.query.get_or_404(id)
    name = group.name
    # 不删除组内学生的历史实训记录，只删除分组本身
    db.session.delete(group)
    db.session.commit()
    flash(f'实训分组 "{name}" 已删除')
    return redirect(url_for('training_groups'))


@app.route('/training/groups/<int:id>/students', methods=['POST'])
def training_group_students(id):
    group = TrainingGroup.query.get_or_404(id)
    student_ids = request.form.getlist('student_ids')
    # 清空原有成员，重新添加
    group.students = []
    for sid in student_ids:
        try:
            s = Student.query.get(int(sid))
            if s:
                group.students.append(s)
        except (ValueError, TypeError):
            pass
    db.session.commit()
    flash(f'已为 "{group.name}" 更新成员（共 {len(student_ids)} 人）')
    return redirect(url_for('training_groups'))


@app.route('/training/stats')
def training_stats():
    sem_id = get_current_semester_id()
    projects = get_semester_projects()
    students = get_semester_students()
    data = []
    for s in students:
        rq = TrainingRecord.query.filter_by(student_id=s.id)
        if sem_id:
            rq = rq.filter_by(semester_id=sem_id)
        records = rq.all()
        project_stats = {}
        for p in projects:
            project_records = [r for r in records if r.project_id == p.id]
            if project_records:
                avg_score = round(sum(r.score for r in project_records) / len(project_records), 1)
                project_stats[p.name] = {
                    'avg_score': avg_score,
                    'count': len(project_records)
                }
        data.append({
            'student': s,
            'records': project_stats,
        })
    return render_template('training_stats.html',
                          students=data, projects=projects)


# ══════════════════════════════════════════════
# 座位管理
# ══════════════════════════════════════════════


@app.route('/seat')
def seat_view():
    """座位管理主页面"""
    sem_id = get_current_semester_id()
    seats = Seat.query
    if sem_id:
        seats = seats.filter_by(semester_id=sem_id)
    seats = seats.order_by(Seat.row_num, Seat.col_num).all()

    assignments = {}
    assign_q = SeatAssignment.query
    if sem_id:
        assign_q = assign_q.filter_by(semester_id=sem_id)
    for a in assign_q.all():
        assignments[a.seat_id] = a

    students = get_semester_students()

    # 获取未安排的学生
    assigned_student_ids = {a.student_id for a in assignments.values()}
    unassigned_students = [s for s in students if s.id not in assigned_student_ids]

    # 获取行数和列数
    max_rows = max((s.row_num for s in seats), default=0)
    max_cols = max((s.col_num for s in seats), default=0)

    # 构建网格
    grid = []
    for r in range(1, max_rows + 1):
        row = []
        for c in range(1, max_cols + 1):
            seat = next((s for s in seats if s.row_num == r and s.col_num == c), None)
            if seat:
                assignment = assignments.get(seat.id)
                row.append({
                    'seat': seat,
                    'assignment': assignment,
                    'student': assignment.student if assignment else None,
                })
            else:
                row.append(None)
        grid.append(row)

    return render_template('seat.html',
                          grid=grid,
                          max_rows=max_rows,
                          max_cols=max_cols,
                          unassigned_students=unassigned_students,
                          seats=seats)


@app.route('/seat/setup', methods=['POST'])
def seat_setup():
    """初始化座位：批量创建Seat记录"""
    sem_id = get_current_semester_id()
    if not sem_id:
        flash('请先选择学期')
        return redirect(url_for('seat_view'))

    try:
        rows = int(request.form.get('rows', 0))
        cols = int(request.form.get('cols', 0))
    except (ValueError, TypeError):
        flash('请输入有效的行数和列数')
        return redirect(url_for('seat_view'))

    if rows < 1 or cols < 1 or rows > 50 or cols > 50:
        flash('行数和列数应在1-50之间')
        return redirect(url_for('seat_view'))

    # 删除当前学期的所有座位及安排
    SeatAssignment.query.filter_by(semester_id=sem_id).delete()
    Seat.query.filter_by(semester_id=sem_id).delete()
    db.session.commit()

    # 批量创建座位
    for r in range(1, rows + 1):
        for c in range(1, cols + 1):
            seat = Seat(
                name=f'第{r}排第{c}列',
                row_num=r,
                col_num=c,
                semester_id=sem_id
            )
            db.session.add(seat)

    db.session.commit()
    flash(f'已创建 {rows}x{cols} = {rows * cols} 个座位')
    return redirect(url_for('seat_view'))


@app.route('/seat/assign', methods=['POST'])
def seat_assign():
    """安排学生入座"""
    sem_id = get_current_semester_id()
    if not sem_id:
        return jsonify({'ok': False, 'error': '请先选择学期'})

    seat_id = request.form.get('seat_id', type=int)
    student_id = request.form.get('student_id', type=int)

    if not seat_id or not student_id:
        return jsonify({'ok': False, 'error': '参数不完整'})

    # 验证座位属于当前学期
    seat = Seat.query.get(seat_id)
    if not seat or (sem_id and seat.semester_id != sem_id):
        return jsonify({'ok': False, 'error': '座位不存在'})

    # 验证学生属于当前学期
    student = Student.query.get(student_id)
    if not student:
        return jsonify({'ok': False, 'error': '学生不存在'})

    # 清除该学生的其他座位安排（一学生一座）
    SeatAssignment.query.filter_by(student_id=student_id, semester_id=sem_id).delete()

    # 清除该座位的旧安排
    SeatAssignment.query.filter_by(seat_id=seat_id, semester_id=sem_id).delete()

    # 创建新安排
    assignment = SeatAssignment(
        seat_id=seat_id,
        student_id=student_id,
        semester_id=sem_id
    )
    db.session.add(assignment)
    db.session.commit()

    return jsonify({'ok': True, 'student_name': student.name})


@app.route('/seat/swap', methods=['POST'])
def seat_swap():
    """拖动交换座位"""
    sem_id = get_current_semester_id()
    if not sem_id:
        return jsonify({'ok': False, 'error': '请先选择学期'})

    data = request.get_json()
    if not data:
        return jsonify({'ok': False, 'error': '无数据'})

    seat_id_1 = data.get('seat_id_1')
    seat_id_2 = data.get('seat_id_2')

    if not seat_id_1 or not seat_id_2:
        return jsonify({'ok': False, 'error': '参数不完整'})

    if seat_id_1 == seat_id_2:
        return jsonify({'ok': False, 'error': '不能与自己交换'})

    # 获取两个座位的当前安排
    a1 = SeatAssignment.query.filter_by(seat_id=seat_id_1, semester_id=sem_id).first()
    a2 = SeatAssignment.query.filter_by(seat_id=seat_id_2, semester_id=sem_id).first()

    if not a1 and not a2:
        return jsonify({'ok': True, 'swapped': False, 'message': '两个座位均空'})

    if a1 and a2:
        # 互换学生
        temp_student_id = a1.student_id
        a1.student_id = a2.student_id
        a2.student_id = temp_student_id
        db.session.commit()
        return jsonify({
            'ok': True, 'swapped': True,
            'student1_name': a1.student.name if a1.student else None,
            'student2_name': a2.student.name if a2.student else None,
        })
    elif a1 and not a2:
        # 座位2空 → 把座位1的学生移到座位2，清空座位1
        a2 = SeatAssignment(seat_id=seat_id_2, student_id=a1.student_id, semester_id=sem_id)
        db.session.add(a2)
        db.session.delete(a1)
        db.session.commit()
        return jsonify({
            'ok': True, 'swapped': False,
            'student_name': a2.student.name if a2.student else None,
            'from_seat': seat_id_1, 'to_seat': seat_id_2,
        })
    else:
        # 座位1空，座位2有人 → 移到座位1
        a1 = SeatAssignment(seat_id=seat_id_1, student_id=a2.student_id, semester_id=sem_id)
        db.session.add(a1)
        db.session.delete(a2)
        db.session.commit()
        return jsonify({
            'ok': True, 'swapped': False,
            'student_name': a1.student.name if a1.student else None,
            'from_seat': seat_id_2, 'to_seat': seat_id_1,
        })


@app.route('/seat/clear', methods=['POST'])
def seat_clear():
    """清空所有座位安排"""
    sem_id = get_current_semester_id()
    if sem_id:
        SeatAssignment.query.filter_by(semester_id=sem_id).delete()
        db.session.commit()
        flash('所有座位安排已清空')
    return redirect(url_for('seat_view'))


@app.route('/seat/auto', methods=['POST'])
def seat_auto():
    """自动排座：随机安排所有未安排的学生"""
    sem_id = get_current_semester_id()
    if not sem_id:
        flash('请先选择学期')
        return redirect(url_for('seat_view'))

    # 获取所有空座位
    seats = Seat.query.filter_by(semester_id=sem_id).order_by(Seat.row_num, Seat.col_num).all()
    assigned_seat_ids = {a.seat_id for a in SeatAssignment.query.filter_by(semester_id=sem_id).all()}
    empty_seats = [s for s in seats if s.id not in assigned_seat_ids]

    if not empty_seats:
        flash('没有空座位可供安排')
        return redirect(url_for('seat_view'))

    # 获取所有未安排的学生
    students = get_semester_students()
    assigned_student_ids = {a.student_id for a in SeatAssignment.query.filter_by(semester_id=sem_id).all()}
    unassigned = [s for s in students if s.id not in assigned_student_ids]

    if not unassigned:
        flash('所有学生已安排座位')
        return redirect(url_for('seat_view'))

    # 随机打乱并分配
    random.shuffle(unassigned)
    random.shuffle(empty_seats)

    count = 0
    for student, seat in zip(unassigned, empty_seats):
        assignment = SeatAssignment(
            seat_id=seat.id,
            student_id=student.id,
            semester_id=sem_id
        )
        db.session.add(assignment)
        count += 1

    db.session.commit()
    flash(f'自动排座完成：已安排 {count} 名学生')
    return redirect(url_for('seat_view'))


# ══════════════════════════════════════════════
# 课表管理
# ══════════════════════════════════════════════

@app.route('/schedule')
def schedule_view():
    schedule_data = get_semester_schedule()
    week = {d: [] for d in range(5)}
    for s in schedule_data:
        if s.day_of_week in week:
            week[s.day_of_week].append(s)
    schedule_images = ScheduleImage.query.order_by(ScheduleImage.created_at.desc()).limit(10).all()

    # 每门课程分配唯一颜色
    all_courses = list(set(s.course_name for s in schedule_data))
    palette = [
        '#e6f7ff', '#f6ffed', '#fff7e6', '#fce4ec', '#f3e5f5',
        '#e0f2f1', '#fff8e1', '#fbe9e7', '#e8eaf6', '#f1f8e9',
        '#e0f7fa', '#fff3e0', '#fce4ec', '#e8eaf6', '#fff9c4',
        '#e0f2f1', '#fff8e1', '#fbe9e7', '#f1f8e9', '#e0f7fa',
        '#ffecb3', '#c8e6c9', '#bbdefb', '#f8bbd0', '#d1c4e9',
        '#b2dfdb', '#ffe0b2', '#ffccbc', '#c5cae9', '#dcedc8',
        '#b2ebf2', '#ffe0b2', '#f8bbd0', '#d1c4e9', '#b2dfdb',
    ]
    course_color = {}
    for i, c in enumerate(sorted(set(all_courses))):
        course_color[c] = palette[i % len(palette)]
    
    schedule_ocr = []
    for img in schedule_images:
        if img.ocr_result:
            try:
                schedule_ocr.append(json.loads(img.ocr_result))
            except (json.JSONDecodeError, TypeError):
                schedule_ocr.append({})
        else:
            schedule_ocr.append({})

    return render_template('schedule.html',
                          week=week, day_names=DAY_NAMES, max_period=MAX_PERIOD,
                          schedule_images=schedule_images,
                          schedule_ocr=schedule_ocr,
                          period_times=PERIOD_TIMES,
                          course_color=course_color)


@app.route('/schedule/add', methods=['POST'])
def schedule_add():
    day_of_week = request.form.get('day_of_week', '0')
    period = request.form.get('period', '1')
    course_name = request.form.get('course_name', '').strip()
    teacher = request.form.get('teacher', '').strip()
    location = request.form.get('location', '').strip()
    is_training = request.form.get('is_training') == 'on'

    if not course_name:
        flash('请输入课程名称')
        return redirect(url_for('schedule_view'))

    try:
        day_of_week = int(day_of_week)
        period = int(period)
    except ValueError:
        flash('参数错误')
        return redirect(url_for('schedule_view'))

    # 检查是否有冲突
    existing = Schedule.query.filter_by(day_of_week=day_of_week, period=period).first()
    if existing:
        flash(f'{DAY_NAMES[day_of_week]} 第{period}节已有课程: {existing.course_name}')
        return redirect(url_for('schedule_view'))

    s = Schedule(
        day_of_week=day_of_week, period=period,
        course_name=course_name, teacher=teacher,
        location=location, is_training=is_training,
        semester_id=get_current_semester_id()
    )
    db.session.add(s)
    db.session.commit()
    flash(f'已添加: {DAY_NAMES[day_of_week]} 第{period}节 {course_name}')
    return redirect(url_for('schedule_view'))


@app.route('/schedule/<int:id>/delete')
def schedule_delete(id):
    s = Schedule.query.get_or_404(id)
    db.session.delete(s)
    db.session.commit()
    flash('课程已删除')
    return redirect(url_for('schedule_view'))


@app.route('/schedule/move', methods=['POST'])
def schedule_move():
    """拖拽调课 - 与目标位置交换（如果目标已占用）"""
    data = request.get_json()
    if not data:
        return jsonify({'ok': False, 'error': '无数据'})

    schedule_id = data.get('id')
    new_day = data.get('day_of_week')
    new_period = data.get('period')

    if not all([schedule_id, new_day is not None, new_period]):
        return jsonify({'ok': False, 'error': '参数不完整'})

    s = Schedule.query.get(schedule_id)
    if not s:
        return jsonify({'ok': False, 'error': '课程不存在'})

    old_day = s.day_of_week
    old_period = s.period

    # 如果目标位置已有课程，交换
    target = Schedule.query.filter_by(day_of_week=new_day, period=new_period).first()
    if target and target.id != s.id:
        target.day_of_week = old_day
        target.period = old_period

    s.day_of_week = new_day
    s.period = new_period
    db.session.commit()
    return jsonify({'ok': True})


@app.route('/schedule/upload', methods=['POST'])
def schedule_upload():
    """上传课表图片（OCR识别）或Excel文件"""
    file = request.files.get('schedule_image')
    if not file:
        flash('请选择文件')
        return redirect(url_for('schedule_view'))

    filename = file.filename.lower()
    is_excel = filename.endswith('.xls') or filename.endswith('.xlsx')
    is_heic = filename.endswith('.heic') or filename.endswith('.heif')
    ext = filename.rsplit('.', 1)[-1] if '.' in filename else 'jpg'

    if is_excel:
        # Excel处理
        temp_path = os.path.join(UPLOAD_FOLDER, f'_schedule_excel_{random.randint(1000,9999)}.{ext}')
        file.save(temp_path)

        try:
            # 先获取所有sheet名称
            sheets = []
            if filename.endswith('.xls'):
                import xlrd
                wb = xlrd.open_workbook(temp_path)
                sheets = wb.sheet_names()
            else:
                from openpyxl import load_workbook
                wb = load_workbook(temp_path, read_only=True)
                sheets = wb.sheetnames

            # 如果是多sheet，引导用户选择
            if len(sheets) > 1:
                final_path = os.path.join(UPLOAD_FOLDER, f'schedule_excel_{datetime.now().strftime("%Y%m%d%H%M%S")}_{random.randint(1000,9999)}.{ext}')
                os.rename(temp_path, final_path)
                # 保存到session用于后续解析
                session['excel_path'] = final_path
                return render_template('schedule_sheets.html',
                                      sheets=sheets,
                                      file_path=final_path,
                                      now=datetime.now())
            else:
                # 直接解析
                courses = parse_excel_schedule(temp_path, sheets[0])
                added = 0
                for c in courses:
                    existing = Schedule.query.filter_by(
                        day_of_week=c['day_of_week'], period=c['period']
                    ).first()
                    if not existing:
                        db.session.add(Schedule(
                            day_of_week=c['day_of_week'],
                            period=c['period'],
                            course_name=c['course_name'],
                            teacher=c.get('teacher', ''),
                            location=c.get('location', ''),
                            is_training=c.get('is_training', False)
                        ))
                        added += 1
                db.session.commit()
                flash(f'已从Excel解析并添加 {added} 条课程')
                try:
                    os.remove(temp_path)
                except:
                    pass
                return redirect(url_for('schedule_view'))
        except Exception as e:
            flash(f'Excel解析失败: {e}')
            try:
                os.remove(temp_path)
            except:
                pass
            return redirect(url_for('schedule_view'))

    # ── 图片处理 ──
    safe_name = f'schedule_{datetime.now().strftime("%Y%m%d%H%M%S")}_{random.randint(1000,9999)}.{ext}'
    save_path = os.path.join(UPLOAD_FOLDER, safe_name)
    file.save(save_path)

    # HEIC格式 → 转成JPEG
    if is_heic:
        try:
            from pillow_heif import register_heif_opener
            register_heif_opener()
            from PIL import Image
            img = Image.open(save_path)
            jpg_name = safe_name.rsplit('.', 1)[0] + '.jpg'
            jpg_path = os.path.join(UPLOAD_FOLDER, jpg_name)
            img.convert('RGB').save(jpg_path, 'JPEG', quality=90)
            os.remove(save_path)
            save_path = jpg_path
            image_path = f'uploads/{jpg_name}'
        except ImportError:
            flash('请安装 pillow-heif 以支持HEIC格式')
            try:
                os.remove(save_path)
            except:
                pass
            return redirect(url_for('schedule_view'))
    else:
        image_path = f'uploads/{safe_name}'

    # ── OCR识别 ──
    import pytesseract
    from PIL import Image, ImageOps

    img = Image.open(save_path)

    try:
        img = ImageOps.exif_transpose(img)
    except Exception:
        pass

    if img.mode != 'RGB':
        img = img.convert('RGB')

    raw_text = pytesseract.image_to_string(img, lang='chi_sim+eng')
    lines = [l.strip() for l in raw_text.split('\n') if l.strip()]

    course_pattern = re.compile(r'[\u4e00-\u9fff]{2,}')
    courses_found = []
    teachers_found = []
    time_slots = []

    for line in lines:
        chinese_phrases = course_pattern.findall(line)
        for phrase in chinese_phrases:
            if len(phrase) >= 4:
                course_kw = ['汽车', '新能源', '构造', '检修', '原理', '驱动', '电气', '底盘', '发动机',
                            '变速箱', '电路', '诊断', '维护', '保养', '检测', '故障', '维修',
                            '语文', '数学', '英语', '体育', '政治', '历史', '物理', '化学', '生物',
                            '音乐', '美术', '心理', '职业', '德育', '实训', '机械', '电子', '编程']
                if any(kw in phrase for kw in course_kw) and len(phrase) <= 30:
                    cleaned = phrase.replace(' ', '')
                    if cleaned not in courses_found:
                        courses_found.append(cleaned)
                elif phrase not in courses_found:
                    common_words = {'新能源汽车', '电气系统', '混合动力', '驱动系统', '工作原理'}
                    if phrase not in common_words:
                        teachers_found.append(phrase)
            elif len(phrase) >= 2 and phrase not in courses_found and phrase not in teachers_found:
                skip_words = {'上午', '下午', '中午', '早上', '晚上', '星期', '周一', '周二', '周三', '周四', '周五',
                            '教室', '班级', '课程', '时间'}
                if phrase not in skip_words and len(phrase) <= 4:
                    teachers_found.append(phrase)

        time_match = re.search(r'(\d{1,2}):(\d{2})\s*[-—–]\s*(\d{1,2}):(\d{2})', line)
        if time_match:
            time_slots.append(f'{time_match.group(1)}:{time_match.group(2)}-{time_match.group(3)}:{time_match.group(4)}')

    ocr_result = json.dumps({
        'courses': courses_found[:20],
        'teachers': list(set(t for t in teachers_found if t not in courses_found))[:10],
        'time_slots': list(set(time_slots))[:10]
    }, ensure_ascii=False)

    img_record = ScheduleImage(image_path=image_path, raw_text=raw_text, ocr_result=ocr_result)
    db.session.add(img_record)

    # 自动填入课表
    added = 0
    period_order = [(d, p) for d in range(5) for p in range(1, MAX_PERIOD + 1)]
    idx = 0
    for c in courses_found:
        skip_keywords = ['榆中县', '职业技术学校', '课程表', '教室', '新能源班']
        if any(kw in c for kw in skip_keywords) or len(c) < 4:
            continue
        if idx < len(period_order):
            d, p = period_order[idx]
            existing = Schedule.query.filter_by(day_of_week=d, period=p).first()
            if not existing:
                db.session.add(Schedule(day_of_week=d, period=p, course_name=c,
                                       semester_id=get_current_semester_id()))
                added += 1
            idx += 1

    db.session.commit()

    msg = f'课表图片已上传'
    if added > 0:
        msg += f'，自动识别并填入 {added} 条课程'
    if courses_found:
        msg += f' 识别到 {len(courses_found)} 门课程'
    if teachers_found:
        msg += f'、{len(set(teachers_found))} 位教师'
    flash(msg)

    return redirect(url_for('schedule_view'))


@app.route('/schedule/parse_sheet', methods=['POST'])
def schedule_parse_sheet():
    """解析指定sheet的Excel课表"""
    file_path = request.form.get('file_path', session.get('excel_path', ''))
    sheet_name = request.form.get('sheet_name', '')

    if not file_path or not os.path.exists(file_path):
        flash('文件不存在或已过期，请重新上传')
        session.pop('excel_path', None)
        return redirect(url_for('schedule_view'))

    try:
        # 先清空当前学期课表
        sem_id = get_current_semester_id()
        Schedule.query.filter_by(semester_id=sem_id).delete() if sem_id else None
        ScheduleImage.query.delete()
        
        courses = parse_excel_schedule(file_path, sheet_name)
        added = 0
        for c in courses:
            db.session.add(Schedule(
                day_of_week=c['day_of_week'],
                period=c['period'],
                course_name=c['course_name'],
                teacher=c.get('teacher', ''),
                location=c.get('location', ''),
                is_training=c.get('is_training', False),
                semester_id=sem_id
            ))
            added += 1
        db.session.commit()
        flash(f'已从「{sheet_name}」导入 {added} 条课程')
    except Exception as e:
        flash(f'解析失败: {e}')
    finally:
        try:
            os.remove(file_path)
        except:
            pass

    return redirect(url_for('schedule_view'))


# ══════════════════════════════════════════════
# 导出功能
# ══════════════════════════════════════════════

def _export_excel_generic(headers, rows, filename):
    """通用Excel导出"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = 'Sheet1'

    # 表头
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font_white = Font(bold=True, size=11, color='FFFFFF')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    # 数据
    for ri, row in enumerate(rows, 2):
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # 列宽自适应
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_length + 2, 8), 40)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    as_attachment=True, download_name=filename, max_age=0)


def _export_pdf_generic(title, headers, rows, filename):
    """通用PDF导出（支持中文）"""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                     Paragraph, Spacer, PageBreak)
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont

    # 注册中文字体
    chinese_font = None
    for font_path in [
        '/System/Library/Fonts/PingFang.ttc',  # macOS PingFang
        '/System/Library/Fonts/STHeiti Light.ttc',
        '/System/Library/Fonts/Helvetica.ttc',
        '/usr/share/fonts/truetype/wqy/wqy-zenhei.ttc',  # Linux
        '/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc',
    ]:
        if os.path.exists(font_path):
            try:
                pdfmetrics.registerFont(TTFont('ChineseFont', font_path))
                chinese_font = 'ChineseFont'
                break
            except:
                continue
    
    # 若找不到中文字体，fallback 到 Helvetica
    font_name = chinese_font if chinese_font else 'Helvetica'
    font_bold = chinese_font if chinese_font else 'Helvetica-Bold'

    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=landscape(A4),
                           topMargin=15*mm, bottomMargin=15*mm,
                           leftMargin=10*mm, rightMargin=10*mm)

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('Title2', parent=styles['Title'],
                                 fontSize=16, leading=20, spaceAfter=10,
                                 fontName=font_name)

    elements = []
    elements.append(Paragraph(title, title_style))
    elements.append(Spacer(1, 10))

    # 准备数据
    table_data = [headers]
    for row in rows:
        table_data.append([str(v) if v is not None else '' for v in row])

    col_count = len(headers)
    page_width = landscape(A4)[0] - 20*mm
    col_width = max(40, page_width / col_count)

    table = Table(table_data, colWidths=[col_width] * col_count, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), font_bold),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('FONTNAME', (0, 1), (-1, -1), font_name),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F2F7FB')]),
    ]))
    elements.append(table)
    doc.build(elements)
    output.seek(0)
    return send_file(output, mimetype='application/pdf',
                    as_attachment=True, download_name=filename, max_age=0)


@app.route('/export/attendance')
def export_attendance():
    fmt = request.args.get('format', 'excel')
    detail = request.args.get('detail', '')
    filter_student = request.args.get('student', '').strip()
    filter_start = request.args.get('start', '')
    filter_end = request.args.get('end', '')
    
    # 构建基础查询
    base_query = Attendance.query
    if filter_start:
        try: base_query = base_query.filter(Attendance.date >= date.fromisoformat(filter_start))
        except: pass
    if filter_end:
        try: base_query = base_query.filter(Attendance.date <= date.fromisoformat(filter_end))
        except: pass
    
    if filter_student:
        qs = Student.query.filter(Student.name.contains(filter_student))
        qs = qs.filter_by(semester_id=get_current_semester_id())
        students = qs.all()
        if students:
            ids = [s.id for s in students]
            base_query = base_query.filter(Attendance.student_id.in_(ids))
        else:
            students = []
    else:
        students = get_semester_students()
    
    if detail:
        headers = ['日期', '姓名', '状态', '原因', '课程', '课时', '凭证图片']
        status_names = {'present': '出勤', 'late': '迟到', 'sick': '病假', 'personal': '事假', 'truant': '旷课'}
        rows = []
        for s in students:
            records = base_query.filter_by(student_id=s.id).order_by(Attendance.date).all()
            for r in records:
                if r.status == 'present':
                    continue
                rows.append([r.date.isoformat(), s.name, status_names.get(r.status, r.status),
                            r.reason or '', r.course_name or '', popcount(r.period) or 1, r.image_path or ''])
    else:
        headers = ['姓名', '病假(节)', '事假(节)', '旷课(节)', '迟到(次)', '总缺勤', '原因摘要', '凭证数']
        rows = []
        for s in students:
            records = base_query.filter_by(student_id=s.id).all()
            sick = sum(popcount(r.period) for r in records if r.status == 'sick')
            personal = sum(popcount(r.period) for r in records if r.status == 'personal')
            truant = sum(popcount(r.period) for r in records if r.status == 'truant')
            late = sum(1 for r in records if r.status == 'late')
            total = sick + personal + truant
            reasons = '; '.join(set(r.reason for r in records if r.status != 'present' and r.reason))[:100]
            images = len([r for r in records if r.image_path])
            rows.append([s.name, sick, personal, truant, late, total, reasons, images])

    if fmt == 'pdf':
        if detail:
            # 带图片的PDF导出
            buf = _export_attendance_detail_pdf(rows, headers)
        else:
            buf = _export_pdf_generic('考勤统计报表', headers, rows, '考勤统计.pdf')
    else:
        if detail:
            buf = _export_attendance_detail_excel(rows, headers)
        else:
            buf = _export_excel_generic(headers, rows, '考勤统计.xlsx')
    return buf


def _export_attendance_detail_pdf(rows, headers):
    """考勤明细PDF导出（含图片，每页最多2张）"""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer, Image as RLImage,
                                    Table, TableStyle, PageBreak)
    from reportlab.lib import colors
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    
    chinese_font = 'Helvetica'
    for fp in ['/System/Library/Fonts/PingFang.ttc', '/System/Library/Fonts/STHeiti Light.ttc']:
        if os.path.exists(fp):
            try:
                pdfmetrics.registerFont(TTFont('CnFont', fp))
                chinese_font = 'CnFont'
                break
            except:
                continue
    
    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=A4,
                           topMargin=15*mm, bottomMargin=15*mm,
                           leftMargin=15*mm, rightMargin=15*mm)
    
    style = ParagraphStyle('N', fontName=chinese_font, fontSize=10, leading=14)
    
    elements = []
    img_count = 0
    for row in rows:
        date_str, name, status, reason, course, period, img_path = row[:7]
        elements.append(Paragraph(
            f'<b>{date_str}</b> {name} | {status} | {course or "-"} 课时{period} | {reason or ""}',
            style))
        elements.append(Spacer(1, 4))
        
        if img_path:
            full_path = os.path.join(BASE_DIR, 'static', img_path)
            if os.path.exists(full_path):
                try:
                    img = RLImage(full_path, width=400, height=300, kind='proportional')
                    elements.append(img)
                    img_count += 1
                except:
                    elements.append(Paragraph(f'[图片: {img_path}]', style))
            elements.append(Spacer(1, 4))
        
        if img_count >= 2:
            elements.append(PageBreak())
            img_count = 0
    
    doc.build(elements)
    output.seek(0)
    return send_file(output, mimetype='application/pdf', as_attachment=True,
                    download_name=f'考勤明细_{date.today().isoformat()}.pdf')


def _export_attendance_detail_excel(rows, headers):
    """考勤明细Excel导出（含嵌入图片，格式同处分导出）"""
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.styles import Font, Alignment, Border, Side
    
    wb = Workbook()
    ws = wb.active
    ws.title = '考勤明细'
    
    headers = ['日期', '姓名', '状态', '原因', '课程', '课时', '图片']
    hf = Font(bold=True, size=11)
    thin = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
    align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = hf; c.border = thin; c.alignment = align
    
    for ri, row in enumerate(rows, 2):
        date_str, name, status, reason, course, period, img_path = row[:7]
        ws.cell(row=ri, column=1, value=date_str).border = thin
        ws.cell(row=ri, column=2, value=name).border = thin
        ws.cell(row=ri, column=3, value=status).border = thin
        ws.cell(row=ri, column=4, value=reason).border = thin
        ws.cell(row=ri, column=5, value=course or '').border = thin
        ws.cell(row=ri, column=6, value=period).border = thin
        
        if img_path:
            full = os.path.join(BASE_DIR, 'static', img_path)
            if os.path.exists(full):
                try:
                    img = XLImage(full)
                    aspect = img.width / img.height if img.height else 1
                    img.width = min(300, img.width)
                    img.height = img.width / aspect
                    ws.add_image(img, f'G{ri}')
                    ws.column_dimensions['G'].width = max(25, img.width * 0.15)
                except:
                    pass
        ws.row_dimensions[ri].height = max(60, 130)
    
    for ci, w in enumerate([14, 12, 10, 22, 16, 8], 1):
        ws.column_dimensions[chr(64 + ci)].width = w
    ws.column_dimensions['G'].width = 30
    
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf, as_attachment=True, download_name=f'考勤明细_{date.today().isoformat()}.xlsx')


@app.route('/export/grades')
def export_grades():
    fmt = request.args.get('format', 'excel')
    subject = request.args.get('subject', '')
    current_subject = subject if subject else request.form.get('subject', '')
    subjects = get_subjects()
    if not current_subject or current_subject not in subjects:
        current_subject = subjects[0] if subjects else ''
    
    students = get_semester_students()
    
    # 按模板格式导出：标题行 → 学期行 → 科目行 → 评分类型行 → 数据行
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    wb = Workbook()
    ws = wb.active
    ws.title = f'{current_subject}评价'
    
    # 样式
    title_font = Font(name='黑体', size=16, bold=True)
    header_font = Font(name='黑体', size=11, bold=True)
    normal_font = Font(name='仿宋', size=10)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
    header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    
    # 标题行
    ws.merge_cells('A1:F1')
    ws.cell(row=1, column=1, value='榆中县职业技术学校学生学习综合评价表').font = title_font
    ws.cell(row=1, column=1).alignment = center
    ws.row_dimensions[1].height = 35
    
    # 学期行
    ws.merge_cells('A2:F2')
    ws.cell(row=2, column=1, value='2025-2026-2 学期').font = Font(name='黑体', size=12)
    ws.cell(row=2, column=1).alignment = center
    ws.row_dimensions[2].height = 22
    
    # 科目行
    ws.merge_cells('A3:F3')
    ws.cell(row=3, column=1, value=f'科目：{current_subject}').font = header_font
    ws.cell(row=3, column=1).alignment = center
    
    # 评分类型行
    score_headers = ['序号', '姓名', '课堂表现', '作业质量', '课堂笔记', '考试成绩', '学科综合成绩']
    for ci, h in enumerate(score_headers, 1):
        c = ws.cell(row=4, column=ci, value=h)
        c.font = header_font; c.fill = header_fill; c.alignment = center; c.border = thin
    
    # 数据行
    for ri, s in enumerate(students, 5):
        g = Grade.query.filter_by(student_id=s.id, subject=current_subject).first()
        ws.cell(row=ri, column=1, value=ri-4).font = normal_font
        ws.cell(row=ri, column=1).alignment = center
        ws.cell(row=ri, column=1).border = thin
        ws.cell(row=ri, column=2, value=s.name).font = normal_font
        ws.cell(row=ri, column=2).alignment = center
        ws.cell(row=ri, column=2).border = thin
        if g:
            for ci, val in enumerate([g.performance_score, g.homework_score, g.notes_score, g.exam_score, g.comprehensive_score], 3):
                c = ws.cell(row=ri, column=ci, value=val)
                c.font = normal_font; c.alignment = center; c.border = thin
        else:
            for ci in range(3, 8):
                ws.cell(row=ri, column=ci).border = thin
    
    # 列宽
    for ci, w in enumerate([8, 12, 12, 12, 12, 12, 14], 1):
        ws.column_dimensions[chr(64 + ci) if ci < 27 else 'A'].width = w
    
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    
    if fmt == 'pdf':
        buf = _export_pdf_generic(f'{current_subject}综合评价', score_headers,
                                 [[ri-4, s.name,
                                   (Grade.query.filter_by(student_id=s.id, subject=current_subject).first().performance_score if Grade.query.filter_by(student_id=s.id, subject=current_subject).first() else 0),
                                   (Grade.query.filter_by(student_id=s.id, subject=current_subject).first().homework_score if Grade.query.filter_by(student_id=s.id, subject=current_subject).first() else 0),
                                   (Grade.query.filter_by(student_id=s.id, subject=current_subject).first().notes_score if Grade.query.filter_by(student_id=s.id, subject=current_subject).first() else 0),
                                   (Grade.query.filter_by(student_id=s.id, subject=current_subject).first().exam_score if Grade.query.filter_by(student_id=s.id, subject=current_subject).first() else 0),
                                   (Grade.query.filter_by(student_id=s.id, subject=current_subject).first().comprehensive_score if Grade.query.filter_by(student_id=s.id, subject=current_subject).first() else '')]
                                  for ri, s in enumerate(students, 5)], f'{current_subject}_评价.pdf')
    
    return send_file(buf, as_attachment=True, download_name=f'{current_subject}_综合评价_{date.today().isoformat()}.xlsx')


# ── 任课管理 ──
# ══════════════════════════════════════════════

@app.route('/teaching')
@login_required
def teaching():
    """任课管理主页面：显示当前学期的科目列表（卡片式）"""
    sem_id = get_current_semester_id()
    q = Subject.query
    if sem_id:
        q = q.filter_by(semester_id=sem_id)
    subjects = q.order_by(Subject.id).all()
    # 每个科目统计学生人数
    subject_data = []
    for sub in subjects:
        count_q = CourseStudent.query.filter_by(course_id=sub.id)
        if sem_id:
            count_q = count_q.filter_by(semester_id=sem_id)
        student_count = count_q.count()
        subject_data.append({
            'subject': sub,
            'student_count': student_count
        })
    return render_template('teaching.html', subject_data=subject_data)


@app.route('/teaching/<int:course_id>/students')
@login_required
def teaching_students(course_id):
    """查看某科目的学生列表"""
    sem_id = get_current_semester_id()
    subject = Subject.query.get_or_404(course_id)
    q = CourseStudent.query.filter_by(course_id=course_id)
    if sem_id:
        q = q.filter_by(semester_id=sem_id)
    students = q.order_by(CourseStudent.id).all()
    return render_template('teaching.html', subject=subject, students=students)


@app.route('/teaching/<int:course_id>/import', methods=['POST'])
@login_required
def teaching_import(course_id):
    """导入学生：文本批量输入 或 Excel导入"""
    sem_id = get_current_semester_id()
    subject = Subject.query.get_or_404(course_id)
    count = 0

    # 方式1：文本框批量导入（每行一个姓名）
    names_text = request.form.get('names', '').strip()
    if names_text:
        names = [n.strip() for n in names_text.split('\n') if n.strip()]
        for name in names:
            # 检查是否已存在该学生
            existing = CourseStudent.query.filter_by(
                name=name, course_id=course_id, semester_id=sem_id
            ).first()
            if not existing:
                cs = CourseStudent(name=name, course_id=course_id,
                                   semester_id=sem_id)
                db.session.add(cs)
                count += 1
        if count:
            db.session.commit()
            flash(f'成功导入 {count} 名学生到「{subject.name}」')
        else:
            flash('未导入任何学生（可能已全部存在）')
        return redirect(url_for('teaching_students', course_id=course_id))

    # 方式2：Excel导入
    file = request.files.get('file')
    if file and file.filename:
        filename = file.filename.lower()
        if not (filename.endswith('.xls') or filename.endswith('.xlsx')):
            flash('请上传Excel文件 (.xls/.xlsx)')
            return redirect(url_for('teaching_students', course_id=course_id))

        import tempfile
        import os
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        file.save(tmp.name)
        tmp.close()
        try:
            names = []
            if filename.endswith('.xls'):
                import xlrd
                wb = xlrd.open_workbook(tmp.name)
                sheet = wb.sheet_by_index(0)
                for r in range(sheet.nrows):
                    val = str(sheet.cell_value(r, 0)).strip()
                    if val:
                        names.append(val)
            else:
                from openpyxl import load_workbook
                wb = load_workbook(tmp.name, read_only=True, data_only=True)
                ws = wb.active
                for row in ws.iter_rows(values_only=True):
                    if row[0] is not None:
                        val = str(row[0]).strip()
                        if val:
                            names.append(val)
                wb.close()
            for name in names:
                existing = CourseStudent.query.filter_by(
                    name=name, course_id=course_id, semester_id=sem_id
                ).first()
                if not existing:
                    cs = CourseStudent(name=name, course_id=course_id,
                                       semester_id=sem_id)
                    db.session.add(cs)
                    count += 1
            if count:
                db.session.commit()
                flash(f'从Excel成功导入 {count} 名学生到「{subject.name}」')
            else:
                flash('Excel中未找到有效学生数据或已全部存在')
        except Exception as e:
            flash(f'Excel解析失败: {e}')
        finally:
            os.unlink(tmp.name)
        return redirect(url_for('teaching_students', course_id=course_id))

    flash('请提供学生姓名（文本框输入）或上传Excel文件')
    return redirect(url_for('teaching_students', course_id=course_id))


@app.route('/teaching/<int:course_id>/student/<int:id>/delete', methods=['POST'])
@login_required
def teaching_student_delete(course_id, id):
    """删除任课科目下的某个学生"""
    cs = CourseStudent.query.get_or_404(id)
    db.session.delete(cs)
    db.session.commit()
    flash('已删除该学生')
    return redirect(url_for('teaching_students', course_id=course_id))


@app.route('/teaching/add-course', methods=['POST'])
@login_required
def teaching_add_course():
    """添加任课科目（名称+教师）"""
    name = request.form.get('name', '').strip()
    teacher = request.form.get('teacher', '').strip()
    class_name = request.form.get('class_name', '').strip()
    if not name:
        flash('请输入科目名称')
        return redirect(url_for('teaching'))
    # 检查当前学期是否已存在同名科目
    sem_id = get_current_semester_id()
    existing = Subject.query.filter_by(name=name, semester_id=sem_id).first()
    if existing:
        flash(f'科目「{name}」已存在')
        return redirect(url_for('teaching'))
    subject = Subject(name=name, teacher=teacher, semester_id=sem_id)
    db.session.add(subject)
    db.session.commit()
    flash(f'科目「{name}」已添加，教师: {teacher or "未设置"}')
    return redirect(url_for('teaching'))


@app.route('/teaching/<int:id>/delete-course', methods=['POST'])
@login_required
def teaching_delete_course(id):
    """删除科目及关联学生"""
    subject = Subject.query.get_or_404(id)
    name = subject.name
    # 删除关联的CourseStudent
    CourseStudent.query.filter_by(course_id=id).delete()
    # 也清理成绩表中的该科目记录
    Grade.query.filter_by(subject=name).delete()
    db.session.delete(subject)
    db.session.commit()
    flash(f'科目「{name}」及其关联学生已删除')
    return redirect(url_for('teaching'))


# ══════════════════════════════════════════════
# ── 科目管理 ──
# ══════════════════════════════════════════════
@app.route('/subject/add', methods=['POST'])
def subject_add():
    name = request.form.get('name', '').strip()
    if not name:
        flash('请输入科目名称')
    elif Subject.query.filter_by(name=name).first():
        flash(f'科目「{name}」已存在')
    else:
        db.session.add(Subject(name=name, teacher=request.form.get('teacher', '').strip(),
                               semester_id=get_current_semester_id()))
        db.session.commit()
        flash(f'科目「{name}」已添加')
    return redirect(url_for('grades'))


@app.route('/subject/<int:id>/delete')
def subject_delete(id):
    s = Subject.query.get_or_404(id)
    name = s.name
    Grade.query.filter_by(subject=name).delete()
    db.session.delete(s)
    db.session.commit()
    flash(f'科目「{name}」已删除')
    return redirect(url_for('grades'))


# ── 实训管理 ──
# ══════════════════════════════════════════════

@app.route('/export/training')
def export_training():
    fmt = request.args.get('format', 'excel')
    projects = get_semester_projects()
    students = get_semester_students()

    headers = ['姓名', '学号']
    for p in projects:
        headers.append(f'{p.name}(平均分)')
        headers.append(f'{p.name}(次数)')
    headers.append('总项目数')

    rows = []
    for s in students:
        records = TrainingRecord.query.filter_by(student_id=s.id).all()
        row = [s.name, s.student_id]
        total_projects = 0
        for p in projects:
            project_records = [r for r in records if r.project_id == p.id]
            if project_records:
                avg = round(sum(r.score for r in project_records) / len(project_records), 1)
                row.append(avg)
                row.append(len(project_records))
                total_projects += 1
            else:
                row.append('')
                row.append(0)
        row.append(total_projects)
        rows.append(row)

    if fmt == 'pdf':
        return _export_pdf_generic('实训统计报表', headers, rows, '实训统计.pdf')
    return _export_excel_generic(headers, rows, '实训统计.xlsx')


@app.route('/export/schedule')
def export_schedule():
    fmt = request.args.get('format', 'excel')
    headers = ['节次', '周一', '周二', '周三', '周四', '周五']
    rows = []
    schedule_data = get_semester_schedule()
    week = {d: {} for d in range(5)}
    for s in schedule_data:
        if s.day_of_week in week:
            week[s.day_of_week][s.period] = s

    for p in range(1, MAX_PERIOD + 1):
        row = [f'第{p}节']
        for d in range(5):
            s = week[d].get(p)
            if s:
                parts = [s.course_name]
                if s.teacher:
                    parts.append(s.teacher)
                if s.location:
                    parts.append(s.location)
                row.append(' / '.join(parts))
            else:
                row.append('')
        rows.append(row)

    if fmt == 'pdf':
        return _export_pdf_generic('班级课表', headers, rows, '班级课表.pdf')
    return _export_excel_generic(headers, rows, '班级课表.xlsx')


# ══════════════════════════════════════════════
# 数据管理（全量导出+恢复+加密备份）
# ══════════════════════════════════════════════

BACKUP_DIR = os.path.join(BASE_DIR, 'static', 'backups')
os.makedirs(BACKUP_DIR, exist_ok=True)

# 当前数据库 schema 版本号（每修改表结构时递增）
SCHEMA_VERSION = '1.0.36'


def _get_db_path():
    """获取当前用户的数据库文件路径"""
    try:
        uid = session.get('user_id')
        if uid:
            return os.path.join(USER_DB_DIR, f'u{uid}.db')
    except:
        pass
    trim_pkgvar = os.environ.get('TRIM_PKGVAR', '')
    if trim_pkgvar:
        return os.path.join(trim_pkgvar, 'master.db')
    return MASTER_DB_PATH


def _get_db_size():
    """获取数据库文件大小（人类可读）"""
    path = _get_db_path()
    if not os.path.exists(path):
        return '文件不存在'
    size = os.path.getsize(path)
    for unit in ['B', 'KB', 'MB', 'GB']:
        if size < 1024:
            return f'{size:.1f} {unit}'
        size /= 1024
    return f'{size:.1f} GB'


# ── 加密备份辅助函数 ──

def _get_user_key(user_id):
    """从用户的密码哈希派生 Fernet 密钥"""
    user = MasterUser.query.get(user_id)
    if not user:
        return None
    # 用 password_hash 的 SHA-256 作为密钥材料，base64 编码后作为 Fernet 密钥
    key = urlsafe_b64encode(hashlib.sha256(user.password_hash.encode()).digest())
    return key


def _encrypt_backup(data: bytes, user_id: int) -> bytes:
    """加密备份数据，返回 CMB1 格式的字节流"""
    key = _get_user_key(user_id)
    f = Fernet(key)
    # 头部格式: CMB1|user_id|timestamp|schema_version
    header = f'CMB1|{user_id}|{datetime.now().isoformat()}|{SCHEMA_VERSION}\n'.encode()
    encrypted = f.encrypt(data)
    return header + encrypted


def _decrypt_backup(data: bytes, user_id: int) -> bytes:
    """解密 CMB1 备份数据，验证用户ID是否匹配"""
    header_end = data.index(b'\n')
    header = data[:header_end].decode()
    parts = header.split('|')
    if parts[0] != 'CMB1':
        raise ValueError('不支持的备份格式')
    file_user_id = int(parts[1])
    # file_timestamp = parts[2]
    file_schema_version = parts[3] if len(parts) >= 4 else ''
    if file_user_id != user_id:
        raise ValueError('此备份不属于当前用户')
    encrypted = data[header_end + 1:]
    key = _get_user_key(user_id)
    f = Fernet(key)
    return f.decrypt(encrypted), file_schema_version


def _parse_cmb_header(data: bytes):
    """解析 CMB 文件头部，返回 dict 或 None"""
    try:
        header_end = data.index(b'\n')
        header = data[:header_end].decode()
        parts = header.split('|')
        return {
            'magic': parts[0],
            'user_id': int(parts[1]),
            'timestamp': parts[2],
            'schema_version': parts[3] if len(parts) >= 4 else '',
        }
    except (ValueError, IndexError, UnicodeDecodeError):
        return None


def _migrate_db_schema(db_path: str, from_version: str = ''):
    """前向兼容：对旧版本的备份数据库执行 schema 迁移（ALTER TABLE 增加缺失列）"""
    try:
        import sqlite3
        conn = sqlite3.connect(db_path)
        c = conn.cursor()
        # 获取已存在的列
        existing = {row[1] for row in c.execute('PRAGMA table_info(student)').fetchall()}
        # 需要迁移的列
        cols_to_add = [
            ('dormitory', 'VARCHAR(32)', ''),
            ('special_family', 'VARCHAR(64)', ''),
            ('special_family_note', 'TEXT', ''),
            ('special_physical', 'VARCHAR(8)', ''),
            ('special_physical_note', 'TEXT', ''),
            ('remark', 'TEXT', ''),
            ('status', 'VARCHAR(16)', 'active'),
            ('withdrawn_reason', 'TEXT', ''),
        ]
        for col, typ, default in cols_to_add:
            if col not in existing:
                try:
                    c.execute(f'ALTER TABLE student ADD COLUMN {col} {typ} DEFAULT "{default}"')
                except:
                    pass
        # subject 表
        try:
            existing_subj = {row[1] for row in c.execute('PRAGMA table_info(subject)').fetchall()}
            if 'class_name' not in existing_subj:
                c.execute('ALTER TABLE subject ADD COLUMN class_name VARCHAR(64) DEFAULT ""')
        except:
            pass
        conn.commit()
        conn.close()
    except:
        pass


@app.route('/data')
@login_required
def data_management():
    """数据管理主页面"""
    stats = {
        'db_size': _get_db_size(),
        'student_count': Student.query.count(),
        'semester_count': Semester.query.count(),
        'attendance_count': Attendance.query.count(),
        'grade_count': Grade.query.count(),
        'discipline_count': Discipline.query.count(),
        'violation_count': ViolationRecord.query.count(),
        'fund_count': ClassFund.query.count(),
        'training_record_count': TrainingRecord.query.count(),
        'schedule_count': Schedule.query.count(),
        'subject_count': Subject.query.count(),
    }

    current_user_id = session.get('user_id')
    backups = []
    if os.path.exists(BACKUP_DIR):
        for fname in sorted(os.listdir(BACKUP_DIR), reverse=True):
            fpath = os.path.join(BACKUP_DIR, fname)
            if not os.path.isfile(fpath):
                continue
            mtime = datetime.fromtimestamp(os.path.getmtime(fpath))
            size = os.path.getsize(fpath)
            size_str = ''
            for unit in ['B', 'KB', 'MB', 'GB']:
                if size < 1024:
                    size_str = f'{size:.1f} {unit}'
                    break
                size /= 1024

            backup_item = {
                'name': fname,
                'size': size_str,
                'mtime': mtime.strftime('%Y-%m-%d %H:%M:%S'),
                'version': '-',
                'user_id': '-',
            }

            # .cmb 文件：解析头部获取版本号和用户信息
            if fname.endswith('.cmb'):
                try:
                    with open(fpath, 'rb') as f:
                        head = _parse_cmb_header(f.read(512))
                    if head:
                        backup_item['version'] = head['schema_version']
                        backup_item['user_id'] = head['user_id']
                except:
                    pass
            # .db 文件：传统格式
            elif fname.endswith('.db'):
                backup_item['version'] = '旧版'
                backup_item['user_id'] = '—'

            # 只显示当前用户的备份（或传统 .db 文件）
            if fname.endswith('.cmb') and backup_item['user_id'] != current_user_id:
                continue

            backups.append(backup_item)

    return render_template('data.html', stats=stats, backups=backups)


@app.route('/data/export')
@login_required
def data_export():
    """全量导出 - 下载原始数据库文件"""
    db_path = _get_db_path()
    if not os.path.exists(db_path):
        flash('数据库文件不存在')
        return redirect(url_for('data_management'))
    return send_file(
        db_path,
        mimetype='application/octet-stream',
        as_attachment=True,
        download_name='class_manager.db',
        max_age=0
    )


@app.route('/data/export-backup', methods=['POST'])
@login_required
def data_export_backup():
    """生成加密的 CMB 格式备份文件到 static/backups/"""
    db_path = _get_db_path()
    if not os.path.exists(db_path):
        flash('数据库文件不存在')
        return redirect(url_for('data_management'))

    user_id = session.get('user_id')

    try:
        import shutil

        # 先复制一份临时 .db 文件
        tmp_path = os.path.join(BACKUP_DIR, f'_tmp_{user_id}.db')
        shutil.copy2(db_path, tmp_path)

        # 读取数据库内容
        with open(tmp_path, 'rb') as f:
            db_data = f.read()
        os.remove(tmp_path)

        # 加密
        encrypted_data = _encrypt_backup(db_data, user_id)

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        backup_name = f'backup_{timestamp}_u{user_id}.cmb'
        backup_path = os.path.join(BACKUP_DIR, backup_name)

        with open(backup_path, 'wb') as f:
            f.write(encrypted_data)

        flash(f'加密备份已创建：{backup_name}（Schema 版本: {SCHEMA_VERSION}）')
    except Exception as e:
        flash(f'备份失败：{e}')

    return redirect(url_for('data_management'))


@app.route('/data/upload', methods=['POST'])
@login_required
def data_upload():
    """上传 .cmb 加密备份或 .db 文件恢复数据"""
    file = request.files.get('backup_file')
    password_hash = request.form.get('password_hash', '').strip()

    if not file:
        flash('请选择要上传的备份文件')
        return redirect(url_for('data_management'))

    user_id = session.get('user_id')
    db_path = _get_db_path()

    try:
        # 自动备份当前数据库
        if os.path.exists(db_path):
            import shutil
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            auto_backup_path = os.path.join(BACKUP_DIR, f'auto_backup_before_restore_{timestamp}.db')
            shutil.copy2(db_path, auto_backup_path)

        file_bytes = file.read()

        if file.filename.endswith('.cmb'):
            # 加密 CMB 格式 — 需要密码
            if not password_hash:
                flash('恢复加密备份需要输入当前登录密码')
                return redirect(url_for('data_management'))

            # 用输入的密码哈希解密（先验证用户身份 — 切换到主数据库查询）
            old_uri = app.config.get('SQLALCHEMY_DATABASE_URI', '')
            switch_db(MASTER_DB_URI)
            user = MasterUser.query.get(user_id)
            if not user or not user.check_password(password_hash):
                flash('密码验证失败，无法解密备份')
                if old_uri != MASTER_DB_URI:
                    switch_db(old_uri)
                return redirect(url_for('data_management'))
            if old_uri != MASTER_DB_URI:
                switch_db(old_uri)

            # 解密
            decrypted_data, schema_version = _decrypt_backup(file_bytes, user_id)

            # 关闭连接
            db.session.remove()
            db.engine.dispose()

            # 写入数据库文件
            with open(db_path, 'wb') as f:
                f.write(decrypted_data)

            # 如果备份来自旧版本，执行 schema 迁移
            if schema_version and schema_version != SCHEMA_VERSION:
                _migrate_db_schema(db_path, schema_version)
                flash(f'数据已恢复！（Schema 从 {schema_version} 迁移至 {SCHEMA_VERSION}）')
            else:
                flash('加密备份已恢复！')
        elif file.filename.endswith('.db'):
            # 传统 .db 格式 — 直接恢复
            db.session.remove()
            db.engine.dispose()
            with open(db_path, 'wb') as f:
                f.write(file_bytes)
            flash('数据已恢复！（传统 .db 格式）')
        else:
            flash('请上传 .cmb 或 .db 格式的文件')
            return redirect(url_for('data_management'))

    except ValueError as e:
        flash(f'恢复失败：{e}')
    except Exception as e:
        flash(f'恢复失败：{e}')

    return redirect(url_for('data_management'))


@app.route('/data/backups')
@login_required
def data_backups():
    """列出当前用户的所有备份文件（JSON）"""
    user_id = session.get('user_id')
    backups = []
    if os.path.exists(BACKUP_DIR):
        for fname in sorted(os.listdir(BACKUP_DIR), reverse=True):
            fpath = os.path.join(BACKUP_DIR, fname)
            if not os.path.isfile(fpath):
                continue
            mtime = datetime.fromtimestamp(os.path.getmtime(fpath))
            size = os.path.getsize(fpath)
            # .cmb 文件：验证用户归属
            if fname.endswith('.cmb'):
                try:
                    with open(fpath, 'rb') as f:
                        head = _parse_cmb_header(f.read(512))
                    if head and head['user_id'] == user_id:
                        backups.append({
                            'name': fname,
                            'size': size,
                            'mtime': mtime.isoformat(),
                            'version': head['schema_version'],
                        })
                except:
                    pass
            elif fname.endswith('.db'):
                backups.append({
                    'name': fname,
                    'size': size,
                    'mtime': mtime.isoformat(),
                    'version': '旧版',
                })
    return jsonify(backups)


@app.route('/data/backups/<filename>/download')
@login_required
def data_backup_download(filename):
    """下载指定的备份文件"""
    safe_name = os.path.basename(filename)
    backup_path = os.path.join(BACKUP_DIR, safe_name)
    if not os.path.exists(backup_path):
        flash('备份文件不存在')
        return redirect(url_for('data_management'))
    return send_file(
        backup_path,
        mimetype='application/octet-stream',
        as_attachment=True,
        download_name=safe_name,
        max_age=0
    )


@app.route('/data/backups/<filename>/delete', methods=['POST'])
@login_required
def data_backup_delete(filename):
    """删除指定的备份文件"""
    safe_name = os.path.basename(filename)
    backup_path = os.path.join(BACKUP_DIR, safe_name)
    if os.path.exists(backup_path):
        try:
            os.remove(backup_path)
            flash(f'备份文件 {safe_name} 已删除')
        except Exception as e:
            flash(f'删除失败：{e}')
    else:
        flash('备份文件不存在')
    return redirect(url_for('data_management'))


@app.route('/data/restore/<filename>', methods=['POST'])
@login_required
def data_restore_from_backup(filename):
    """从服务器上的已有 .cmb 备份恢复数据"""
    safe_name = os.path.basename(filename)
    password = request.form.get('password', '').strip()

    if not safe_name.endswith('.cmb'):
        flash('仅支持从 .cmb 加密备份恢复')
        return redirect(url_for('data_management'))

    backup_path = os.path.join(BACKUP_DIR, safe_name)
    if not os.path.exists(backup_path):
        flash('备份文件不存在')
        return redirect(url_for('data_management'))

    user_id = session.get('user_id')
    db_path = _get_db_path()

    try:
        # 验证密码
        user = MasterUser.query.get(user_id)
        if not user or not user.check_password(password):
            flash('密码错误，无法解密备份')
            return redirect(url_for('data_management'))

        # 读取并解密备份
        with open(backup_path, 'rb') as f:
            file_bytes = f.read()
        decrypted_data, schema_version = _decrypt_backup(file_bytes, user_id)

        # 自动备份当前数据库
        if os.path.exists(db_path):
            import shutil
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            auto_backup_path = os.path.join(BACKUP_DIR, f'auto_backup_before_restore_{timestamp}.db')
            shutil.copy2(db_path, auto_backup_path)

        # 关闭连接并替换数据库
        db.session.remove()
        db.engine.dispose()
        with open(db_path, 'wb') as f:
            f.write(decrypted_data)

        # 前向兼容：schema 迁移
        if schema_version and schema_version != SCHEMA_VERSION:
            _migrate_db_schema(db_path, schema_version)
            flash(f'数据已从「{safe_name}」恢复！（Schema 从 {schema_version} 迁移至 {SCHEMA_VERSION}）')
        else:
            flash(f'数据已从「{safe_name}」恢复！')

    except ValueError as e:
        flash(f'恢复失败：{e}')
    except Exception as e:
        flash(f'恢复失败：{e}')

    return redirect(url_for('data_management'))


# ── 旧数据库迁移辅助函数 ──
def _run_alter_migrations(db_file):
    """对指定的数据库文件执行ALTER TABLE迁移（新增列）"""
    if not os.path.exists(db_file):
        return
    import sqlite3
    try:
        conn = sqlite3.connect(db_file)
        c = conn.cursor()
        for col, typ in [('dormitory','VARCHAR(32)'),('special_family','VARCHAR(64)'),
                         ('special_family_note','TEXT'),('special_physical','VARCHAR(8)'),
                         ('special_physical_note','TEXT'),('remark','TEXT')]:
            try: c.execute(f'ALTER TABLE student ADD COLUMN {col} {typ} DEFAULT ""')
            except: pass
        for col, typ in [('status','VARCHAR(16)')]:
            try: c.execute(f'ALTER TABLE student ADD COLUMN {col} {typ} DEFAULT "active"')
            except: pass
        for col, typ in [('withdrawn_reason','TEXT')]:
            try: c.execute(f'ALTER TABLE student ADD COLUMN {col} {typ} DEFAULT ""')
            except: pass
        try: c.execute('ALTER TABLE subject ADD COLUMN class_name VARCHAR(64) DEFAULT ""')
        except: pass
        conn.commit(); conn.close()
    except:
        pass


def _do_migration(old_db_path):
    """从旧 class_manager.db 迁移数据到新的多用户数据库结构"""
    import sqlite3

    try:
        old_conn = sqlite3.connect(old_db_path)
        old_c = old_conn.cursor()
    except Exception as e:
        print(f'❌ 无法打开旧数据库: {e}')
        return

    # 检查旧数据库是否有 user 表
    old_c.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='user'")
    if not old_c.fetchone():
        print('⚠️ 旧数据库没有 user 表，跳过用户迁移')
        old_conn.close()
        return

    # 读取旧用户
    old_c.execute('SELECT id, username, password_hash FROM "user" ORDER BY id')
    old_users = old_c.fetchall()
    migrated_user_ids = set()

    for uid, uname, pw_hash in old_users:
        # 在master.db中创建MasterUser
        if not MasterUser.query.filter_by(username=uname).first():
            mu = MasterUser(username=uname, password_hash=pw_hash if pw_hash else '')
            if not mu.password_hash:
                mu.set_password('123456')
            db.session.add(mu)
            db.session.commit()
        else:
            mu = MasterUser.query.filter_by(username=uname).first()

        migrated_user_ids.add(mu.id)

    print(f'📋 迁移了 {len(old_users)} 个用户到主数据库')

    if not migrated_user_ids:
        old_conn.close()
        return

    # 第一个用户（通常是admin，id最小的）接收所有旧业务数据
    admin_id = sorted(migrated_user_ids)[0]
    old_tables = []
    old_c.execute("SELECT name FROM sqlite_master WHERE type='table'")
    old_tables = [r[0] for r in old_c.fetchall() if r[0] not in ('user', 'sqlite_sequence', 'master_user')]

    if not old_tables:
        old_conn.close()
        print('⚠️ 旧数据库没有业务表，跳过数据迁移')
        # 重命名旧数据库以防重复迁移
        _rename_old_db(old_db_path)
        return

    # 切换到admin的业务数据库
    switch_to_user_db(admin_id)

    # 先检查新db中是否有数据，避免重复迁移
    for tbl in old_tables:
        try:
            old_c.execute(f'SELECT COUNT(*) FROM "{tbl}"')
            old_count = old_c.fetchone()[0]
        except:
            old_count = 0

        new_cursor = db.session.execute(f'SELECT COUNT(*) FROM "{tbl}"') if old_count > 0 else None
        new_count = 0
        if new_cursor:
            try:
                new_count = new_cursor.fetchone()[0]
            except:
                pass

        if old_count > 0 and new_count == 0:
            # 读取所有数据
            table_info = []
            try:
                old_c.execute(f'PRAGMA table_info("{tbl}")')
                table_info = old_c.fetchall()
            except:
                continue

            col_names = [col[1] for col in table_info if col[1]]
            if not col_names:
                continue

            # 读取旧数据
            try:
                old_c.execute(f'SELECT * FROM "{tbl}"')
                rows = old_c.fetchall()
            except:
                continue

            # 逐行插入
            placeholders = ','.join(['?' for _ in col_names])
            for row in rows:
                try:
                    db.session.execute(
                        f'INSERT INTO "{tbl}" ({",".join(col_names)}) VALUES ({placeholders})',
                        row
                    )
                except Exception as insert_err:
                    pass  # 跳过冲突行

            print(f'  📄 迁移 {tbl}: {len(rows)} 条')

    db.session.commit()
    old_conn.close()

    # 执行ALTER TABLE迁移
    user_db_path = os.path.join(BASE_DIR, 'instance', 'users', f'u{admin_id}.db')
    _run_alter_migrations(user_db_path)

    # 重命名旧数据库以防重复迁移
    _rename_old_db(old_db_path)

    # 切回主数据库
    switch_db(MASTER_DB_URI)


def _rename_old_db(old_db_path):
    """重命名旧数据库文件，标记为已迁移"""
    try:
        import shutil
        if os.path.exists(old_db_path):
            backup_name = old_db_path + '.migrated'
            if not os.path.exists(backup_name):
                shutil.move(old_db_path, backup_name)
                print(f'📦 旧数据库已备份为: {backup_name}')
            else:
                os.remove(old_db_path)
    except:
        pass


# ══════════════════════════════════════════════
# 启动
# ══════════════════════════════════════════════

if __name__ == '__main__':
    with app.app_context():
        # ── 第1步：初始化主数据库（master.db） ──
        switch_db(MASTER_DB_URI)
        print('✅ 主数据库已就绪')

        # ── 第2步：检查旧数据库，执行数据迁移 ──
        old_db_path = os.path.join(BASE_DIR, 'instance', 'class_manager.db')
        if os.path.exists(old_db_path):
            print('📦 检测到旧数据库，正在迁移数据...')
            _do_migration(old_db_path)
            print('✅ 旧数据库数据已迁移完成')

        # ── 第3步：创建默认管理员（首次运行） ──
        if not MasterUser.query.filter_by(username='admin').first():
            admin = MasterUser(username='admin')
            admin.set_password('admin123')
            db.session.add(admin)
            db.session.commit()
            print('已创建默认管理员: admin / admin123')
            # 为admin创建业务数据库并初始化
            switch_to_user_db(admin.id)
            from datetime import date
            if Semester.query.count() == 0:
                s = Semester(name='2025-2026学年度第1学期', start_date=date(2025,9,1),
                             end_date=date(2026,1,15), is_current=True)
                db.session.add(s)
                db.session.commit()
                print('已创建默认学期')
            # 切回主数据库（启动后 before_request 会处理）
            switch_db(MASTER_DB_URI)
        else:
            # 确保所有已有用户都有业务数据库
            for mu in MasterUser.query.all():
                user_db_path = os.path.join(BASE_DIR, 'instance', 'users', f'u{mu.id}.db')
                if not os.path.exists(user_db_path):
                    switch_to_user_db(mu.id)
                    # 执行ALTER TABLE迁移
                    _run_alter_migrations(user_db_path)
                    from datetime import date
                    if Semester.query.count() == 0:
                        s = Semester(name='2025-2026学年度第1学期', start_date=date(2025,9,1),
                                     end_date=date(2026,1,15), is_current=True)
                        db.session.add(s)
                        db.session.commit()
                        print(f'已为用户 {mu.username} 创建业务数据库和默认学期')
                else:
                    # 对已有数据库执行ALTER TABLE迁移
                    _run_alter_migrations(user_db_path)
            switch_db(MASTER_DB_URI)

    print(f'✅ 班级管理系统启动成功！')
    print(f'🌐 请访问: http://localhost:5800')
    print(f'🔑 管理员账号: admin / admin123\n')
    import sys
    port = 5800
    for i, a in enumerate(sys.argv[1:], 1):
        if a == '--port' and i < len(sys.argv[1:])+1:
            try: port = int(sys.argv[i+1])
            except: pass
    # 使用 waitress 生产级服务器（反代兼容）
    try:
        from waitress import serve
        print(f'🚀 使用 waitress 服务器 (生产模式)')
        serve(app, host='0.0.0.0', port=port, threads=16)
    except ImportError:
        # 回退 Flask 开发服务器（带 threaded=True）
        print(f'⚠️ 使用 Flask 开发服务器（建议安装 waitress: pip install waitress）')
        app.run(host='0.0.0.0', port=port, threaded=True, debug=False)
